#!/usr/bin/env python3
"""
merge_urls.py  —  Fold an image-URL or compliance-doc file into existing products.

Many vendors deliver images / testing docs in a separate spreadsheet keyed by DS
number (after they've been uploaded to Cloudinary). This merges those URLs into the
matching product JSON files.

Auto-detects the file type from its columns:
  - "Image URL 1..N" columns        -> product.images.urls
  - "Doc N Type" / "Doc N URL" pairs -> product.compliance.documents [{type,url}]

NON-DESTRUCTIVE by default (the failsafe): if a product already has images/docs that
differ from the incoming set, it is reported as a CONFLICT and skipped. Pass
--overwrite to replace. DS numbers with no matching product are reported as unmatched.

Usage:
  python scripts/merge_urls.py "C:\\path\\CanaltoysURLs.xlsx"
  python scripts/merge_urls.py "file.xlsx" --dry-run
  python scripts/merge_urls.py "file.xlsx" --overwrite
"""

import argparse
import glob
import json
import os
import sys

for _stream in (sys.stdout, sys.stderr):
    try:
        _stream.reconfigure(encoding="utf-8")
    except (AttributeError, ValueError):
        pass

try:
    from openpyxl import load_workbook
except ImportError:
    sys.exit("ERROR: openpyxl not installed. Run: pip install openpyxl")

REPO_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
PRODUCTS_DIR = os.path.join(REPO_ROOT, "products")


def is_http(v):
    return isinstance(v, str) and v.strip().lower().startswith(("http://", "https://"))


def build_ds_index():
    """ds_number (str) -> product file path, across all vendor folders."""
    idx = {}
    for path in glob.glob(os.path.join(PRODUCTS_DIR, "**", "*.json"), recursive=True):
        with open(path, encoding="utf-8") as f:
            p = json.load(f)
        ds = str(p.get("ds_number") or "").strip()
        if ds:
            idx[ds] = path
    return idx


def find_header_row(ws, max_scan=12):
    """Find the row index whose cells contain a 'DS Number' header."""
    for r in range(1, min(max_scan, ws.max_row) + 1):
        for c in range(1, ws.max_column + 1):
            val = ws.cell(r, c).value
            if isinstance(val, str) and val.strip().lower() in ("ds number", "ds#", "ds number "):
                return r
    return None


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("xlsx")
    ap.add_argument("--overwrite", action="store_true",
                    help="replace existing images/docs even if they differ")
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()

    path = args.xlsx.strip().strip('"').strip("'")
    if not os.path.isfile(path):
        sys.exit(f"ERROR: file not found -> {path}")

    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    hdr_row = find_header_row(ws)
    if not hdr_row:
        sys.exit("ERROR: could not find a 'DS Number' header row in the first sheet.")

    headers = {c: (ws.cell(hdr_row, c).value or "") for c in range(1, ws.max_column + 1)}
    ds_col = next((c for c, h in headers.items() if str(h).strip().lower() in ("ds number", "ds#")), None)
    img_cols = [c for c, h in headers.items() if str(h).strip().lower().startswith("image url")]
    doc_type_cols = {}  # doc index -> type col
    doc_url_cols = {}   # doc index -> url col
    for c, h in headers.items():
        hl = str(h).strip().lower()
        if hl.startswith("doc ") and hl.endswith(" type"):
            doc_type_cols[hl.replace("type", "").strip()] = c
        elif hl.startswith("doc ") and hl.endswith(" url"):
            doc_url_cols[hl.replace("url", "").strip()] = c

    mode = "images" if img_cols else ("docs" if doc_url_cols else None)
    if not mode:
        sys.exit("ERROR: file has neither 'Image URL' nor 'Doc N URL' columns.")
    print(f"Detected mode: {mode}   (header row {hdr_row})")

    idx = build_ds_index()
    matched, unmatched, conflicts, updated = [], [], [], []

    for r in range(hdr_row + 1, ws.max_row + 1):
        ds_raw = ws.cell(r, ds_col).value
        if ds_raw is None or str(ds_raw).strip() == "":
            continue
        ds = str(ds_raw).strip()
        if ds.endswith(".0"):
            ds = ds[:-2]

        if mode == "images":
            incoming = [ws.cell(r, c).value for c in img_cols]
            incoming = [u.strip() for u in incoming if is_http(u)]
        else:
            incoming = []
            for key in sorted(doc_url_cols):
                url = ws.cell(r, doc_url_cols[key]).value
                if is_http(url):
                    dtype = ws.cell(r, doc_type_cols[key]).value if key in doc_type_cols else None
                    incoming.append({"type": (str(dtype).strip() if dtype else None), "url": url.strip()})

        if not incoming:
            continue

        if ds not in idx:
            unmatched.append(ds)
            continue
        matched.append(ds)

        with open(idx[ds], encoding="utf-8") as f:
            prod = json.load(f)

        if mode == "images":
            existing = (prod.get("images") or {}).get("urls") or []
            target_desc = "images.urls"
        else:
            existing = (prod.get("compliance") or {}).get("documents") or []
            target_desc = "compliance.documents"

        if existing and existing != incoming and not args.overwrite:
            conflicts.append(ds)
            print(f"  CONFLICT DS{ds}: {target_desc} already set and differs - skipping "
                  f"(use --overwrite). have={len(existing)} incoming={len(incoming)}")
            continue
        if existing == incoming:
            continue

        if args.dry_run:
            print(f"  ~ would set {len(incoming)} {mode} on DS{ds}")
            updated.append(ds)
            continue

        if mode == "images":
            prod.setdefault("images", {})["urls"] = incoming
        else:
            prod.setdefault("compliance", {})["documents"] = incoming
        with open(idx[ds], "w", encoding="utf-8") as f:
            json.dump(prod, f, indent=2, ensure_ascii=False, default=str)
            f.write("\n")
        updated.append(ds)
        print(f"  set {len(incoming)} {mode} on DS{ds}")

    print("\n" + "=" * 56)
    print(f"  Matched: {len(matched)}   Updated: {len(updated)}   "
          f"Conflicts: {len(conflicts)}   Unmatched DS#: {len(unmatched)}")
    if unmatched:
        print("  Unmatched (no product in catalog): " + ", ".join(unmatched[:20])
              + (" ..." if len(unmatched) > 20 else ""))
    print("  Next: python scripts/validate.py && python scripts/build_index.py")
    print("=" * 56)


if __name__ == "__main__":
    main()
