#!/usr/bin/env python3
"""
import_setup_sheet.py  —  Bring a completed DS setup sheet into the catalog.

Reads the 'DS Only' sheet of a 2026 New Item Setup Sheet (.xlsx) and writes
one JSON file per product into products/<brand-slug>/DS<number>.json.

By default this is NON-DESTRUCTIVE: if a product file already exists it is
SKIPPED (so re-importing never clobbers hand-edits).  Pass --update to
overwrite existing files (the git diff / PR is your review gate).

Usage:
  python scripts/import_setup_sheet.py "path/to/Brand 2026 Setup.xlsx"
  python scripts/import_setup_sheet.py "file.xlsx" --update
  python scripts/import_setup_sheet.py "file.xlsx" --dry-run
"""

import argparse
import json
import os
import sys
from datetime import date

# Product names can contain non-cp1252 chars (Ō, em-dash, …). Keep the Windows
# console from crashing when we print them.
for _stream in (sys.stdout, sys.stderr):
    try:
        _stream.reconfigure(encoding="utf-8")
    except (AttributeError, ValueError):
        pass

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

try:
    from openpyxl import load_workbook
except ImportError:
    sys.exit("ERROR: openpyxl not installed. Run: pip install openpyxl")

from ds_automation import DS, DS_DATA_START
from ds_schema import sheet_row_to_product, brand_slug

REPO_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
PRODUCTS_DIR = os.path.join(REPO_ROOT, "products")


def vendor_from_path(path):
    """
    Derive the vendor name from the file path's '...\\Priority Vendors\\<Vendor>\\...'
    segment. Returns None if not found (caller should then require --vendor).
    """
    parts = os.path.normpath(path).replace("\\", "/").split("/")
    for i, seg in enumerate(parts):
        if seg.strip().lower() == "priority vendors" and i + 1 < len(parts):
            return parts[i + 1].strip()
    return None


def read_supplier_id(ws):
    """DS Only cell B3 holds the Apprise supplier number (or 'xxxx' placeholder)."""
    raw = ws.cell(row=3, column=2).value
    s = str(raw).strip() if raw is not None else ""
    if not s or s.lower() == "xxxx":
        return None
    try:
        return int(s)
    except ValueError:
        return s


def main():
    ap = argparse.ArgumentParser(description="Import a DS setup sheet into the catalog.")
    ap.add_argument("xlsx", help="Path to the .xlsx setup sheet")
    ap.add_argument("--vendor", help="Vendor name (folder). Default: auto-detected from the "
                                     "'Priority Vendors/<Vendor>' path segment.")
    ap.add_argument("--update", action="store_true",
                    help="Overwrite existing product files instead of skipping them")
    ap.add_argument("--dry-run", action="store_true",
                    help="Show what would happen without writing files")
    ap.add_argument("--sheet", help="Which tab to read. Default: 'Move Forward Items' if "
                                    "present (the curated carry-forward list), else 'DS Only'.")
    args = ap.parse_args()

    path = args.xlsx.strip().strip('"').strip("'")
    if not os.path.isfile(path):
        sys.exit(f"ERROR: file not found -> {path}")

    vendor = args.vendor or vendor_from_path(path)
    if not vendor:
        sys.exit("ERROR: could not detect vendor from path. Pass --vendor \"<Vendor Name>\".")
    slug = brand_slug(vendor)
    print(f"Vendor: {vendor}  (folder: products/{slug}/)")

    print(f"Opening: {os.path.basename(path)}")
    wb = load_workbook(path, data_only=True)
    if args.sheet:
        sheet_name = args.sheet
    elif "Move Forward Items" in wb.sheetnames:
        sheet_name = "Move Forward Items"
    elif "DS Only" in wb.sheetnames:
        sheet_name = "DS Only"
    else:
        sys.exit("ERROR: no 'Move Forward Items' or 'DS Only' sheet - is this the 2026 template?")
    if sheet_name not in wb.sheetnames:
        sys.exit(f"ERROR: sheet '{sheet_name}' not found. Tabs: {wb.sheetnames}")
    print(f"Reading sheet: '{sheet_name}'")
    ws = wb[sheet_name]

    supplier_id = read_supplier_id(ws)
    source_file = os.path.basename(path)
    today = date.today().isoformat()

    created, updated, skipped = [], [], []

    for row in range(DS_DATA_START, ws.max_row + 1):
        ds_number = ws.cell(row=row, column=DS.DS_NUMBER).value
        if ds_number is None or str(ds_number).strip() == "":
            continue

        prod = sheet_row_to_product(ws, row)
        prod["vendor"] = vendor
        prod["_meta"]["source_file"] = source_file
        prod["_meta"]["source_sheet"] = sheet_name
        prod["_meta"]["imported_at"] = today
        prod["_meta"]["supplier_id"] = supplier_id
        prod["_meta"]["vendor_slug"] = slug

        ds_clean = str(prod["ds_number"]).strip()
        vendor_dir = os.path.join(PRODUCTS_DIR, slug)
        out_path = os.path.join(vendor_dir, f"DS{ds_clean}.json")
        rel = os.path.relpath(out_path, REPO_ROOT)
        exists = os.path.exists(out_path)

        if exists and not args.update:
            skipped.append(rel)
            print(f"  - skip   DS{ds_clean}  ({rel}) already exists - use --update to overwrite")
            continue

        if args.dry_run:
            (updated if exists else created).append(rel)
            print(f"  ~ would {'update' if exists else 'create'}  DS{ds_clean}  {prod.get('product_name')}")
            continue

        os.makedirs(vendor_dir, exist_ok=True)
        with open(out_path, "w", encoding="utf-8") as f:
            json.dump(prod, f, indent=2, ensure_ascii=False, default=str)
            f.write("\n")
        (updated if exists else created).append(rel)
        print(f"  {'updated' if exists else 'created'}  DS{ds_clean}  {prod.get('product_name')}")

    print("\n" + "=" * 56)
    print(f"  Created: {len(created)}   Updated: {len(updated)}   Skipped: {len(skipped)}")
    if supplier_id is None:
        print("  NOTE: supplier id (B3) was blank/placeholder - set it later if needed.")
    print("  Next: run  python scripts/validate.py  then  python scripts/build_index.py")
    print("=" * 56)


if __name__ == "__main__":
    main()
