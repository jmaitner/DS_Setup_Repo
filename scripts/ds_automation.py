#!/usr/bin/env python3
"""
DS WMS Upload File Generator
======================================================
Reads a completed New Item Setup Sheet (2026 template)
and generates 5 WMS upload files:

  1. New Product      (WMS .xlsx)  ← only NEW items (DS# >= first new DS#)
  2. Supplier         (WMS .xlsx)  ← all items
  3. Product Unit of Measure (UOM)  (WMS .xlsx)  ← all items
  4. Purchase Cost    (WMS .xlsx)  ← all items
  5. Goflow Product Import  (_goflow.csv)  ← all items

After generating the New Product file, a validation pass reads it back
and compares every DS#, name, and UPC against the setup sheet.  Any
mismatch is printed as a table so you can catch problems before uploading.

Optionally generates 7 channel setup files:
  6.  Walmart          (_walmart.xlsx)
  7.  TikTok Shop      (_tiktok.xlsx)
  8.  Best Buy         (_bestbuy.xlsx)
  9.  Toys R Us / Logicbroker  (_toysrus.xlsx)
  10. Target Plus      (_target.csv)
  11. eBay             (_ebay.csv)
  12. Shopify          (_shopify.xlsx)

All data is read exclusively from the DS Only sheet.

Usage:
  python ds_automation.py                    # interactive prompts
  python ds_automation.py "path/to/file.xlsx"  # pass file path as argument

Requirements:
  pip install openpyxl
"""

import sys
import os
import csv
import re
import shutil
import tempfile
from datetime import date

try:
    from openpyxl import load_workbook, Workbook
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)


# ══════════════════════════════════════════════════════════════
#  CONFIGURABLE DEFAULTS  (edit these to change behaviour)
# ══════════════════════════════════════════════════════════════

DEFAULT_LEAD_TIME   = 12            # days
DEFAULT_MOQ         = 12            # minimum order quantity
DEFAULT_CHANNEL     = "DS"          # distribution channel
DEFAULT_MERCH_TYPE  = "Merch"       # merchandise type
PURCHASE_END_DATE   = "12/31/2099"  # end date for purchase cost records
SHORT_NAME_LENGTH   = 15            # max chars for the short product name (col 6)
DISPLAY_NAME_LENGTH = 20            # max chars for the display name (cols 2 & 5)

# DS Only sheet: row where column headers live
DS_HEADER_ROW  = 4
# DS Only sheet: first row of actual product data
DS_DATA_START  = 7


# ══════════════════════════════════════════════════════════════
#  DS ONLY COLUMN MAPPING  (1-based)
#  2026 template — 81 columns total.
#  If a future template update shifts columns, only edit here.
# ══════════════════════════════════════════════════════════════

class DS:
    """DS Only sheet column indices (1-based). Single source of truth."""

    # ── Pricing ───────────────────────────────────────────────
    DS_NUMBER            = 1
    DROP_SHIP_COST       = 2
    WHOLESALE_COST       = 3
    DS_COST_DOMESTIC     = 4
    IMPORT_COST          = 5

    # ── Product identity ──────────────────────────────────────
    VENDOR_ITEM_NUMBER   = 6
    ID_TYPE              = 7   # EAN / UPC / GTIN selector
    UPC                  = 8
    BRAND                = 9
    PRODUCT_NAME         = 10
    MSRP                 = 11
    FOB_POINT            = 12
    COUNTRY_OF_ORIGIN    = 13
    HARMONIZED_CODE      = 14

    # ── Content ───────────────────────────────────────────────
    BULLET_1             = 15
    BULLET_2             = 16
    BULLET_3             = 17
    BULLET_4             = 18
    BULLET_5             = 19
    KEYWORDS             = 20
    DESCRIPTION          = 21

    # ── Attributes ────────────────────────────────────────────
    MATERIAL             = 22
    NUM_PIECES           = 23
    WHATS_IN_BOX         = 24
    PRIMARY_COLOR        = 25
    SECONDARY_COLOR      = 26
    MIN_AGE              = 27
    MAX_AGE              = 28
    GENDER               = 29
    ASSEMBLY_REQUIRED    = 30
    ASSEMBLY_INSTRUCTIONS = 31

    # ── Images ────────────────────────────────────────────────
    IMAGE_AVAILABILITY   = 32
    IMAGE_URL_1          = 33
    IMAGE_URL_2          = 34
    IMAGE_URL_3          = 35
    IMAGE_URL_4          = 36
    IMAGE_URL_5          = 37

    # ── Item dimensions (product inside packaging) ────────────
    ITEM_WEIGHT          = 38
    ITEM_WEIGHT_UNIT     = 39
    ITEM_LENGTH          = 40
    ITEM_LENGTH_UNIT     = 41
    ITEM_WIDTH           = 42
    ITEM_WIDTH_UNIT      = 43
    ITEM_HEIGHT          = 44
    ITEM_HEIGHT_UNIT     = 45

    # ── Package dimensions — source of truth for UOM "each" ───
    PKG_WEIGHT           = 46
    PKG_WEIGHT_UNIT      = 47
    PKG_LENGTH           = 48
    PKG_LENGTH_UNIT      = 49
    PKG_WIDTH            = 50
    PKG_WIDTH_UNIT       = 51
    PKG_HEIGHT           = 52
    PKG_HEIGHT_UNIT      = 53

    # ── Master case — source of truth for UOM "case" ──────────
    CASE_QTY             = 54
    MC_LENGTH            = 55
    MC_LENGTH_UNIT       = 56
    MC_WIDTH             = 57
    MC_WIDTH_UNIT        = 58
    MC_HEIGHT            = 59
    MC_HEIGHT_UNIT       = 60
    MC_WEIGHT            = 61
    MC_WEIGHT_UNIT       = 62

    # ── Compliance / safety ───────────────────────────────────
    CHOKING_HAZARD       = 63
    LEAD_PHTHALATES      = 64
    WARRANTY_INCLUDED    = 65
    WARRANTY_DESC        = 66
    BATTERIES_REQUIRED   = 67
    BATTERIES_INCLUDED   = 68
    BATTERY_CELL_COMP    = 69
    BATTERY_TYPE_QTY     = 70
    PACKAGING_TYPE       = 71

    # ── Compliance documents ──────────────────────────────────
    COMPLIANCE_CERT      = 72
    DOC                  = 73
    SDS                  = 74
    SDS_URL              = 75
    CPSIA                = 76
    TEST_REPORTS         = 77
    CPC                  = 78
    PRODUCT_PICS         = 79
    INSTRUCTIONS         = 80
    LETTER_OF_COMPLIANCE = 81


# ══════════════════════════════════════════════════════════════
#  HELPERS
# ══════════════════════════════════════════════════════════════

def _to_ds_int(ds_number):
    """Safely convert a DS number to int for numeric comparison."""
    try:
        return int(ds_number)
    except (TypeError, ValueError):
        return 0


def case_uom_name(qty):
    """Return UOM name like 'case012' from a quantity of 12."""
    try:
        return f"case{int(qty):03d}"
    except (TypeError, ValueError):
        return "case001"


def cubic_feet(length, width, height):
    """Calculate cubic feet from three dimensions in inches."""
    try:
        result = (float(length) * float(width) * float(height)) / 1728
        return round(result, 4)
    except (TypeError, ValueError):
        return None


def safe_str(value, max_len=None):
    """Convert value to string and optionally truncate."""
    if value is None:
        return ""
    s = str(value).strip()
    if max_len:
        s = s[:max_len]
    return s


def normalize_weight_unit(value):
    """Normalise weight unit strings to what Apprise expects ('lb', 'oz', 'kg', 'g')."""
    mapping = {
        "lbs": "lb", "pounds": "lb", "pound": "lb",
        "ounces": "oz", "ozs": "oz",
        "kilograms": "kg", "kilogram": "kg", "kgs": "kg",
        "grams": "g", "gm": "g",
    }
    return mapping.get(str(value).strip().lower(), str(value).strip()) if value else "lb"


def normalize_dim_unit(value):
    """Normalise dimension unit strings to what Apprise expects ('inch', 'ft', 'cm', 'm')."""
    mapping = {
        "inches": "inch", "in": "inch", "ins": "inch",
        "feet": "ft", "foot": "ft",
        "centimeters": "cm", "centimetre": "cm", "centimetres": "cm",
        "meters": "m", "metre": "m", "metres": "m",
    }
    return mapping.get(str(value).strip().lower(), str(value).strip()) if value else "inch"


def to_grams(weight, unit):
    """Convert a weight value to grams (float) for Shopify's Variant Grams field."""
    try:
        w = float(weight)
    except (TypeError, ValueError):
        return None
    u = str(unit).strip().lower() if unit else "lb"
    if u in ("lb", "lbs"):
        return round(w * 453.592, 4)
    if u == "oz":
        return round(w * 28.3495, 4)
    if u in ("kg", "kgs"):
        return round(w * 1000, 4)
    if u in ("g",):
        return round(w, 4)
    return round(w * 453.592, 4)   # fallback: assume lb


def dim_to_inches(value, unit):
    """Convert a dimension value to inches (float). Returns None if value is blank."""
    try:
        v = float(value)
    except (TypeError, ValueError):
        return None
    u = str(unit).strip().lower() if unit else "inch"
    if u in ("inch", "inches", "in", "ins"):
        return round(v, 4)
    if u in ("ft", "feet", "foot"):
        return round(v * 12, 4)
    if u in ("cm", "centimeter", "centimeters", "centimetre", "centimetres"):
        return round(v / 2.54, 4)
    if u in ("m", "meter", "meters", "metre", "metres"):
        return round(v * 39.3701, 4)
    return round(v, 4)   # fallback: pass through


def shopify_handle(name):
    """Generate a Shopify URL handle from a product name."""
    if not name:
        return ""
    h = str(name).lower()
    h = re.sub(r'[^a-z0-9\s-]', '', h)
    h = re.sub(r'[\s-]+', '-', h)
    return h.strip('-')


def coerce_row(row):
    """Replace any None values in a row with '' so openpyxl writes all cells."""
    return ["" if v is None else v for v in row]


def _safe_save_wb(wb, output_path):
    """Save workbook via a temp file then copy — avoids permission errors on network shares."""
    tmp_fd, tmp_path = tempfile.mkstemp(suffix=".xlsx")
    os.close(tmp_fd)
    try:
        wb.save(tmp_path)
        shutil.copy2(tmp_path, output_path)
    finally:
        if os.path.exists(tmp_path):
            os.remove(tmp_path)


# ══════════════════════════════════════════════════════════════
#  TEMPLATE READER  —  DS Only sheet only
# ══════════════════════════════════════════════════════════════

def read_template(filepath):
    """
    Read the 2026 New Item Setup Sheet and return a tuple of
    (products, supplier_id) where supplier_id is read from DS Only
    cell B3.  supplier_id will be None if the cell is blank or
    contains the placeholder value 'xxxx'.
    All product data is read from the DS Only sheet.
    """
    print(f"\nOpening: {os.path.basename(filepath)}")
    wb = load_workbook(filepath, data_only=True)

    if "DS Only" not in wb.sheetnames:
        raise ValueError("Sheet 'DS Only' not found. Check the file is the correct 2026 template.")

    ws = wb["DS Only"]

    # ── Read Apprise Supplier Number from B3 ──────────────────
    b3_raw = ws.cell(row=3, column=2).value
    b3_str = str(b3_raw).strip() if b3_raw is not None else ""
    if b3_str and b3_str.lower() != "xxxx":
        try:
            sheet_supplier_id = int(b3_str)
        except ValueError:
            sheet_supplier_id = b3_str
    else:
        sheet_supplier_id = None

    def ds(row_idx, col):
        return ws.cell(row=row_idx, column=col).value

    products = []
    max_row = ws.max_row

    for row in range(DS_DATA_START, max_row + 1):
        ds_number = ds(row, DS.DS_NUMBER)
        if ds_number is None or str(ds_number).strip() == "":
            continue  # skip rows without a DS Number

        product = {
            # ── Core identifiers ──────────────────────────────
            "ds_number":           ds_number,
            "vendor_item_number":  ds(row, DS.VENDOR_ITEM_NUMBER),
            "id_type":             ds(row, DS.ID_TYPE),
            "upc":                 ds(row, DS.UPC),
            "brand":               ds(row, DS.BRAND),
            "product_name":        ds(row, DS.PRODUCT_NAME),
            "cost":                ds(row, DS.WHOLESALE_COST),
            "msrp":                ds(row, DS.MSRP),
            "drop_ship_cost":      ds(row, DS.DROP_SHIP_COST),
            "country_of_origin":   ds(row, DS.COUNTRY_OF_ORIGIN),

            # ── Content ───────────────────────────────────────
            "description":         ds(row, DS.DESCRIPTION),
            "bullet1":             ds(row, DS.BULLET_1),
            "bullet2":             ds(row, DS.BULLET_2),
            "bullet3":             ds(row, DS.BULLET_3),
            "bullet4":             ds(row, DS.BULLET_4),
            "bullet5":             ds(row, DS.BULLET_5),
            "keywords":            ds(row, DS.KEYWORDS),

            # ── Attributes ────────────────────────────────────
            "material":            ds(row, DS.MATERIAL),
            "num_pieces":          ds(row, DS.NUM_PIECES),
            "primary_color":       ds(row, DS.PRIMARY_COLOR),
            "min_age":             ds(row, DS.MIN_AGE),
            "max_age":             ds(row, DS.MAX_AGE),
            "assembly_required":   ds(row, DS.ASSEMBLY_REQUIRED),

            # ── Images ────────────────────────────────────────
            "image_url_1":         ds(row, DS.IMAGE_URL_1),
            "image_url_2":         ds(row, DS.IMAGE_URL_2),
            "image_url_3":         ds(row, DS.IMAGE_URL_3),
            "image_url_4":         ds(row, DS.IMAGE_URL_4),
            "image_url_5":         ds(row, DS.IMAGE_URL_5),

            # ── Item dimensions (for Goflow / channel listings)
            "item_weight":         ds(row, DS.ITEM_WEIGHT),
            "item_weight_unit":    ds(row, DS.ITEM_WEIGHT_UNIT),
            "item_length":         ds(row, DS.ITEM_LENGTH),
            "item_width":          ds(row, DS.ITEM_WIDTH),
            "item_height":         ds(row, DS.ITEM_HEIGHT),

            # ── Package dimensions (source of truth for UOM each)
            "pkg_weight":          ds(row, DS.PKG_WEIGHT),
            "pkg_weight_unit":     ds(row, DS.PKG_WEIGHT_UNIT),
            "pkg_length":          ds(row, DS.PKG_LENGTH),
            "pkg_length_unit":     ds(row, DS.PKG_LENGTH_UNIT),
            "pkg_width":           ds(row, DS.PKG_WIDTH),
            "pkg_width_unit":      ds(row, DS.PKG_WIDTH_UNIT),
            "pkg_height":          ds(row, DS.PKG_HEIGHT),
            "pkg_height_unit":     ds(row, DS.PKG_HEIGHT_UNIT),

            # ── Master case (source of truth for UOM case) ────
            "case_qty":            ds(row, DS.CASE_QTY),
            "mc_length":           ds(row, DS.MC_LENGTH),
            "mc_length_unit":      ds(row, DS.MC_LENGTH_UNIT),
            "mc_width":            ds(row, DS.MC_WIDTH),
            "mc_width_unit":       ds(row, DS.MC_WIDTH_UNIT),
            "mc_height":           ds(row, DS.MC_HEIGHT),
            "mc_height_unit":      ds(row, DS.MC_HEIGHT_UNIT),
            "mc_weight":           ds(row, DS.MC_WEIGHT),
            "mc_weight_unit":      ds(row, DS.MC_WEIGHT_UNIT),

            # ── Compliance / safety ───────────────────────────
            "choking_hazard":      ds(row, DS.CHOKING_HAZARD),
            "batteries_required":  ds(row, DS.BATTERIES_REQUIRED),
            "batteries_included":  ds(row, DS.BATTERIES_INCLUDED),
            "battery_cell_comp":   ds(row, DS.BATTERY_CELL_COMP),
            "battery_type_qty":    ds(row, DS.BATTERY_TYPE_QTY),
        }

        missing = [k for k in ("product_name", "upc", "cost", "case_qty")
                   if product.get(k) is None or str(product.get(k, "")).strip() == ""]
        if missing:
            print(f"  ⚠ WARNING: DS#{ds_number} is missing: {', '.join(missing)}")

        products.append(product)
        print(f"  Found product: DS#{ds_number}  |  {product['product_name']}  |  UPC: {product['upc']}")

    return products, sheet_supplier_id


# ══════════════════════════════════════════════════════════════
#  WMS OUTPUT FILE GENERATORS
# ══════════════════════════════════════════════════════════════

def generate_new_product(products, supplier_id, category, output_path, min_ds_number=0):
    """
    New Product upload file.
    248-column format. One row per product.
    Trailing space in col 248 of the last data row (WMS requirement).

    min_ds_number: if > 0, only items with DS# >= this value are included.
    """
    if min_ds_number:
        before = len(products)
        products = [p for p in products if _to_ds_int(p["ds_number"]) >= min_ds_number]
        skipped = before - len(products)
        if skipped:
            print(f"  (Skipped {skipped} existing item(s) with DS# < {min_ds_number})")

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    all_rows = []
    for p in products:
        case_name = case_uom_name(p["case_qty"])
        name      = safe_str(p["product_name"])
        short_name   = name[:SHORT_NAME_LENGTH]
        display_name = name[:DISPLAY_NAME_LENGTH]

        row = [""] * 248

        row[0]   = p["ds_number"]
        row[1]   = display_name
        row[2]   = DEFAULT_CHANNEL
        row[3]   = DEFAULT_MERCH_TYPE
        row[4]   = display_name
        row[5]   = short_name
        row[6]   = p["upc"]
        row[7]   = "each"
        row[8]   = category
        row[21]  = "Default"
        row[32]  = "All"
        row[38]  = "No"
        row[60]  = "Yes"
        row[61]  = "No"
        row[63]  = "No"
        row[64]  = "No"
        row[65]  = "No"
        row[66]  = "Yes"
        row[67]  = "Yes"
        row[68]  = "FIFO/UM/SEQ"
        row[69]  = "Default"
        row[72]  = "lb"
        row[80]  = "Yes"
        row[81]  = "Yes"
        row[92]  = "each"
        row[93]  = "each"
        row[94]  = "each"
        row[98]  = "Yes"
        row[99]  = "Yes"
        row[100] = "No"
        row[101] = "No"
        row[102] = "Yes"
        row[103] = "Yes"
        row[106] = "No"
        row[108] = supplier_id
        row[115] = 1
        row[116] = 1
        row[119] = 1
        row[120] = 1
        row[121] = "Yes"
        row[126] = case_name
        row[131] = "Yes"
        row[141] = "Purchase"
        row[151] = "Yes"
        row[154] = "No"
        row[213] = "No"

        all_rows.append(row)

    if all_rows:
        all_rows[-1][247] = " "

    for row in all_rows:
        ws.append(coerce_row(row))

    wb.save(output_path)
    print(f"  ✓ New Product      → {os.path.basename(output_path)}  ({len(all_rows)} item(s))")


def generate_supplier(products, supplier_id, lead_time, moq, output_path):
    """
    Supplier upload file.
    18-column format. One row per product.
    Trailing space in col 18 of the last data row (WMS requirement).
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    all_rows = []
    for p in products:
        row = [""] * 18
        row[0]  = supplier_id
        row[1]  = p["ds_number"]
        row[2]  = p["vendor_item_number"]
        row[5]  = lead_time
        row[6]  = moq
        row[12] = "Yes"
        row[14] = "each"
        all_rows.append(row)

    if all_rows:
        all_rows[-1][17] = " "

    for row in all_rows:
        ws.append(coerce_row(row))

    wb.save(output_path)
    print(f"  ✓ Supplier         → {os.path.basename(output_path)}")


def generate_uom(products, output_path):
    """
    Product Unit of Measure upload file.
    34-column format. 4 rows per product: P/each, L/each, P/case, L/case.

    Each rows  → package dimensions  (DS Only cols 46-53)
    Case rows  → master case dims    (DS Only cols 54-62)
    Weight/dimension units are read from the sheet; default to lb/inch if blank.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    all_rows = []
    for p in products:
        case_name = case_uom_name(p["case_qty"])

        # Unit values — read from sheet, normalise to Apprise expected strings
        each_wt_unit  = normalize_weight_unit(p["pkg_weight_unit"])
        each_dim_unit = normalize_dim_unit(p["pkg_length_unit"])
        case_wt_unit  = normalize_weight_unit(p["mc_weight_unit"])
        case_dim_unit = normalize_dim_unit(p["mc_length_unit"])

        cf_each = cubic_feet(p["pkg_length"], p["pkg_width"], p["pkg_height"])
        cf_case = cubic_feet(p["mc_length"],  p["mc_width"],  p["mc_height"])

        def make_each_row(type_code):
            r = [""] * 34
            r[0]  = type_code
            r[1]  = p["ds_number"]
            r[2]  = DEFAULT_CHANNEL
            r[4]  = "each"
            r[5]  = "Yes"
            r[6]  = "Yes"
            r[7]  = "Yes"
            r[8]  = "Yes"
            r[9]  = "Yes"
            r[10] = "Yes"
            r[11] = "Yes"
            r[12] = p["pkg_length"]
            r[13] = p["pkg_width"]
            r[14] = p["pkg_height"]
            r[15] = p["pkg_weight"]
            r[17] = each_wt_unit
            r[23] = p["upc"]
            r[27] = "Yes"
            r[29] = each_dim_unit
            r[30] = cf_each
            r[31] = "cuft"
            return r

        def make_case_row(type_code):
            r = [""] * 34
            r[0]  = type_code
            r[1]  = p["ds_number"]
            r[2]  = DEFAULT_CHANNEL
            r[4]  = case_name
            r[5]  = "Yes"
            r[6]  = "Yes"
            r[7]  = "Yes"
            r[8]  = "No"
            r[9]  = "No"
            r[10] = "No"
            r[11] = "No"
            r[12] = p["mc_length"]
            r[13] = p["mc_width"]
            r[14] = p["mc_height"]
            r[15] = p["mc_weight"]
            r[17] = case_wt_unit
            r[24] = "each"
            r[29] = case_dim_unit
            r[30] = cf_case
            r[31] = "cuft"
            return r

        all_rows.append(make_each_row("P"))
        all_rows.append(make_each_row("L"))
        all_rows.append(make_case_row("P"))
        all_rows.append(make_case_row("L"))

    if all_rows:
        all_rows[-1][33] = " "

    for row in all_rows:
        ws.append(coerce_row(row))

    wb.save(output_path)
    print(f"  ✓ Product UOM      → {os.path.basename(output_path)}")


def generate_purchase_cost(products, supplier_id, output_path):
    """
    Purchase Cost upload file.
    79-column format. One row per product.
    Start date = today. End date = 12/31/2099.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    today = date.today()

    for p in products:
        row = [""] * 79
        row[0]  = supplier_id
        row[2]  = p["ds_number"]
        row[5]  = p["cost"]
        row[6]  = p["cost"]
        row[74] = today
        row[75] = PURCHASE_END_DATE
        row[76] = "USD"
        row[77] = "US"
        row[78] = "No"
        ws.append(row)

    wb.save(output_path)
    print(f"  ✓ Purchase Cost    → {os.path.basename(output_path)}")


def generate_goflow_csv(products, output_path):
    """
    Goflow Product Import CSV.
    Writes to a local temp file first, then copies to the destination
    to avoid permission errors on network shares.
    Item dimensions (not package dims) used for L/W/H/weight here
    since Goflow uses these for product listings.
    """
    headers = [
        "Item Number", "Name", "Description", "Brand", "Manufacturer",
        "Condition", "Is Purchasable", "Is Sellable", "Identifier > UPC",
        "Cost", "Price", "MSRP", "Pounds", "Ounces", "Length", "Width", "Height"
    ]

    tmp_fd, tmp_path = tempfile.mkstemp(suffix=".csv")
    try:
        with os.fdopen(tmp_fd, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(headers)
            for p in products:
                writer.writerow([
                    p["ds_number"],
                    safe_str(p["product_name"]),
                    safe_str(p["description"]),
                    safe_str(p["brand"]),
                    safe_str(p["brand"]),
                    "New",
                    "Yes",
                    "Yes",
                    p["upc"],
                    p["cost"],
                    p["msrp"],
                    p["msrp"],
                    p["item_weight"],
                    0,
                    p["item_length"],
                    p["item_width"],
                    p["item_height"],
                ])
        shutil.copy2(tmp_path, output_path)
    finally:
        if os.path.exists(tmp_path):
            os.remove(tmp_path)

    print(f"  ✓ Goflow Import    → {os.path.basename(output_path)}")


# ══════════════════════════════════════════════════════════════
#  VALIDATION
# ══════════════════════════════════════════════════════════════

def validate_new_product(products, output_path, min_ds_number=0):
    """
    Read the generated new_product.xlsx back and compare every row's
    DS#, display name, and UPC against the setup sheet data.
    """
    print("\n  Running pre-upload validation...")

    if min_ds_number:
        expected = [p for p in products if _to_ds_int(p["ds_number"]) >= min_ds_number]
    else:
        expected = list(products)

    try:
        wb = load_workbook(output_path, data_only=True)
        ws = wb.active
    except Exception as e:
        print(f"  ⚠ Could not open output file for validation: {e}")
        return

    file_rows = []
    for row in ws.iter_rows(min_row=1, values_only=True):
        ds_in_file   = str(row[0]).strip() if row[0] is not None else ""
        name_in_file = str(row[1]).strip() if row[1] is not None else ""
        upc_in_file  = str(row[6]).strip() if row[6] is not None else ""
        if ds_in_file:
            file_rows.append((ds_in_file, name_in_file, upc_in_file))

    mismatches = []

    if len(file_rows) != len(expected):
        mismatches.append(
            f"  ROW COUNT: file has {len(file_rows)} rows, "
            f"setup sheet has {len(expected)} new items"
        )

    for i, (ds_in_file, name_in_file, upc_in_file) in enumerate(file_rows):
        if i >= len(expected):
            mismatches.append(
                f"  Row {i+1}: EXTRA row in file  DS#={ds_in_file}  "
                f"name={name_in_file!r}  UPC={upc_in_file}"
            )
            continue

        p = expected[i]
        exp_ds   = str(p["ds_number"]).strip()
        exp_name = safe_str(p["product_name"])[:DISPLAY_NAME_LENGTH]
        exp_upc  = str(p["upc"]).strip() if p["upc"] is not None else ""

        row_issues = []
        if ds_in_file != exp_ds:
            row_issues.append(f"DS# got={ds_in_file!r} expected={exp_ds!r}")
        if name_in_file != exp_name:
            row_issues.append(f"name got={name_in_file!r} expected={exp_name!r}")
        if upc_in_file != exp_upc:
            row_issues.append(f"UPC got={upc_in_file!r} expected={exp_upc!r}")

        if row_issues:
            mismatches.append(f"  Row {i+1}: " + " | ".join(row_issues))

    for i in range(len(file_rows), len(expected)):
        p = expected[i]
        mismatches.append(
            f"  Row {i+1}: MISSING from file  "
            f"DS#={p['ds_number']}  name={safe_str(p['product_name'])[:DISPLAY_NAME_LENGTH]!r}"
        )

    if not mismatches:
        print(f"  ✓ Validation PASSED — all {len(file_rows)} item(s) match the setup sheet")
    else:
        print(f"\n  {'!' * 56}")
        print(f"  !! VALIDATION FAILED — {len(mismatches)} issue(s) found.")
        print(f"  !! DO NOT upload new_product.xlsx to Apprise until resolved.")
        print(f"  {'!' * 56}")
        for line in mismatches:
            print(line)
        print()


# ══════════════════════════════════════════════════════════════
#  CHANNEL SETUP FILE GENERATORS
# ══════════════════════════════════════════════════════════════

# ── Walmart header rows (228 columns) ─────────────────────────
_WM_HEADERS = [
    None, None, None,
    "Product ID Type", "Product ID", None, "SKU", "9-Digit Supplier Id",
    "Product Name", "Brand Name", "Walmart Private Label/Unbranded",
    "Selling Price", "Unit Cost", "Walmart Factory ID (+)",
    "Walmart Factory ID 1 (+)", "Walmart Factory ID 2 (+)",
    "Walmart Factory ID 3 (+)", "Walmart Factory ID 4 (+)",
    "Walmart Factory ID 5 (+)", "Walmart Factory ID 6 (+)",
    "Walmart Factory ID 7 (+)", "Walmart Factory ID 8 (+)",
    "Walmart Factory ID 9 (+)", "Walmart Factory ID 10 (+)",
    "Country of Origin (+)", "Country of Origin 1 (+)",
    "Country of Origin 2 (+)",
    "Product is or Contains an Electronic Component?",
    "Product is or Contains this Battery Type",
    "Shipping Dimensions Depth (in)", "Shipping Dimensions Width (in)",
    "Shipping Dimensions Height (in)", "Shipping Weight (lbs)",
    "Ships in Original Packaging", "Product is or contains a chemical",
    "Product is a pesticide (Includes pesticide devices and antimicrobial treated products)",
    "Product is or contains an aerosol", "Site Description",
    "Additional Image URL (+)", "Additional Image URL 1 (+)",
    "Additional Image URL 10 (+)", "Additional Image URL 11 (+)",
    "Additional Image URL 12 (+)", "Additional Image URL 13 (+)",
    "Additional Image URL 14 (+)", "Additional Image URL 15 (+)",
    "Additional Image URL 16 (+)", "Additional Image URL 17 (+)",
    "Additional Image URL 18 (+)", "Additional Image URL 19 (+)",
    "Additional Image URL 2 (+)", "Additional Image URL 20 (+)",
    "Additional Image URL 21 (+)", "Additional Image URL 22 (+)",
    "Additional Image URL 23 (+)", "Additional Image URL 24 (+)",
    "Additional Image URL 25 (+)", "Additional Image URL 3 (+)",
    "Additional Image URL 4 (+)", "Additional Image URL 5 (+)",
    "Additional Image URL 6 (+)", "Additional Image URL 7 (+)",
    "Additional Image URL 8 (+)", "Additional Image URL 9 (+)",
    "Color", "Color Category (+)", "Color Category 1 (+)",
    "Color Category 2 (+)", "Condition", "Count Per Pack",
    "Has Written Warranty", "Is Prop 65 Warning Required",
    "Key Features (+)", "Key Features 1 (+)", "Key Features 2 (+)",
    "Key Features 3 (+)", "Key Features 4 (+)", "Key Features 5 (+)",
    "Key Features 6 (+)", "Key Features 7 (+)", "Key Features 8 (+)",
    "Main Image URL", "Material (+)", "Material 1 (+)",
    "Measure", "Unit", "Measure", "Unit", "Multipack Quantity", "Unit",
    "Measure", "Small Parts Warning Code (+)", "Small Parts Warning Code 1 (+)",
    "Total Count", "Age Group (+)", "Age Group 1 (+)", "Age Group 2 (+)",
    "Age Group 3 (+)", "Age Group 4 (+)", "Age Group 5 (+)",
    "Measure", "Unit", "Measure", "Unit", "Measure", "Unit",
    "Fill Material (+)", "Has NRTL Listing Certification", "Is Collectible",
    "Number of Pieces", "Slime Type", "Stuffed Animal & Plush Toy Type",
    "Toy Camera Type", "Accessories Included (+)", "Accessories Included 1 (+)",
    "Accessories Included 2 (+)", "Accessories Included 3 (+)",
    "Accessories Included 4 (+)", "Accessories Included 5 (+)",
    "Accessories Included 6 (+)", "Activity (+)", "Activity 1 (+)",
    "Additional Features (+)", "Additional Features 1 (+)",
    "Additional Features 2 (+)", "Additional Features 3 (+)",
    "Animal Type (+)", "Animal Type 1 (+)", "Measure", "Unit",
    "Battery Count", "Battery Size", "Brand License (+)",
    "California Prop 65 Warning Text", "Character (+)", "Character 1 (+)",
    "Character Group (+)", "Character Group 1 (+)", "Edition",
    "Educational Focus (+)", "Educational Focus 1 (+)",
    "Educational Level (+)", "Educational Level 1 (+)",
    "Educational Level 2 (+)", "Educational Level 3 (+)",
    "Educational Level 4 (+)", "Era", "Fabric Material Percentage",
    "Fabric Material Name", "Harmonized System Code Country",
    "Harmonized System Code", "Harmonized System Code Country",
    "Harmonized System Code", "Harmonized System Code Country",
    "Harmonized System Code", "Harmonized System Code Country",
    "Harmonized System Code", "Harmonized System Code Country",
    "Harmonized System Code", "Harmonized System Code Country",
    "Harmonized System Code", "Harmonized System Code Country",
    "Harmonized System Code", "Harmonized System Code Country",
    "Harmonized System Code",
    "Is Food Component Monetary Value Over 50 Percent",
    "Items Included (+)", "Items Included 1 (+)", "Items Included 10 (+)",
    "Items Included 2 (+)", "Items Included 3 (+)", "Items Included 4 (+)",
    "Items Included 5 (+)", "Items Included 6 (+)", "Items Included 7 (+)",
    "Items Included 8 (+)", "Items Included 9 (+)",
    "Law Label Registration Number", "Law Label Identification Provider",
    "Manufacturer Name", "Manufacturer Part Number", "Media Franchise",
    "Model Number", "Net Content Statement", "NRTL Test Standard",
    "NRTL Organization", "Nutrition/Supplement/Drug Label Type",
    "Occasion (+)", "Occasion 1 (+)", "Pattern (+)", "Pattern 1 (+)",
    "Power Type (+)", "Product Line (+)", "Product Line 1 (+)",
    "Resolution", "Size Descriptor", "Sports League (+)", "Sports Team (+)",
    "Texture (+)", "Texture 1 (+)", "Theme",
    "Third Party Accreditation Symbol on Product Package Code (+)",
    "Third Party Accreditation Symbol on Product Package Code 1 (+)",
    "Title", "Warranty Text", "Warranty URL", "Variant Group ID",
    "Variant Attribute Names (+)", "Variant Attribute Names 1 (+)",
    "Is Primary Variant", "Swatch Variant Attribute", "Swatch Image URL",
    "Is Preorder", "Release Date", "Site Start Date", "Site End Date",
    "Value", "Type", "Batch Number Indicator", "Minimum Advertised Price",
    "Must ship alone?", "Packaging Marked Returnable Indicator",
    "ZIP Codes", "States", "State Restrictions Reason",
    "Sustainability Feature Code (+)", "Pesticide Type",
    "Suggested Reuse Item Number",
]

# ── TikTok template metadata rows ─────────────────────────────
_TIKTOK_ROW1 = [
    "category", "brand", "product_name", "product_description", "main_image",
    "image_2", "image_3", "image_4", "image_5", "image_6", "image_7",
    "image_8", "image_9", "gtin_type", "gtin_code",
    "property_name_1", "property_value_1",
    "property_1_image", "property_1_image_2", "property_1_image_3",
    "property_1_image_4", "property_1_image_5", "property_1_image_6",
    "property_1_image_7", "property_1_image_8", "property_1_image_9",
    "property_name_2", "property_value_2",
    "parcel_weight", "parcel_length", "parcel_width", "parcel_height",
    "delivery", "price", "list_price", "quantity", "seller_sku",
    "size_chart", "special_product_listing_type",
    "product_property/100107", "product_property/100177",
    "product_property/100336", "product_property/100433",
    "product_property/100443", "product_property/100492",
    "product_property/100495", "product_property/100496",
    "product_property/100706", "product_property/100707",
    "product_property/100708", "product_property/101200",
    "product_property/101611", "product_property/101614",
    "product_property/101619", "product_property/101623",
    "product_property/101624", "product_property/101625",
    "product_property/102256", "product_property/101395",
    "product_property/101398", "product_property/101400",
    "product_property/101397", "product_property/101610",
    "product_property/100216",
    "qualification/1729439947062478079",
    "qualification/1729439947134305535",
    "qualification/1729439947134502143",
    "qualification/8647636475739801353",
    "aimed_product_status",
]
_TIKTOK_ROW2 = [
    "V5.0.0", "create_product", "imperial", "category_v2", None,
    "021775741762426fdbdfdbdfdbdfdbd000000000000008ff56d13",
    "normal_file", "sale_platforms",
] + [None] * 61
_TIKTOK_ROW3 = [
    "Category", "Brand", "Product name", "Product description", "Main image",
    "Image 2", "Image 3", "Image 4", "Image 5", "Image 6", "Image 7",
    "Product Image 8", "Product Image 9",
    "Identifier Code Type", "Identifier Code",
    "Primary variation name (theme)", "Primary variation value (option)",
    "Primary variation image 1", "Primary variation image 2",
    "Primary variation image 3", "Primary variation image 4",
    "Primary variation image 5", "Primary variation image 6",
    "Primary variation image 7", "Primary variation image 8",
    "Primary variation image 9",
    "Secondary variation name (theme)", "Secondary variation value (option)",
    "Package weight(lb)", "Package length(inch)", "Package width(inch)",
    "Package height(inch)", "Delivery options",
    "Retail Price (Local Currency)", "List price (Local currency)",
    "Quantity", "Seller SKU", "Size Chart", "Auction product",
    "Warranty Type", "Model", "Region Of Origin", "Recommended Age",
    "Feature", "Manufacturer", "Warranty Duration", "Material Feature",
    "Manufacturer", "Interest", "Cartoon Characters", "CPSIA Tracking Label",
    "How Batteries Are Packed", "Battery Or Cell Weight In Grams",
    "Dangerous Goods Or Hazardous Materials", "Number Of Batteries Or Cells",
    "Battery Or Cell Capacity In Wh", "Battery Or Cell Capacity In Grams",
    "Other Dangerous Goods Or Hazardous Materials",
    "CA Prop 65: Repro. Chems", "Reprotoxic Chemicals",
    "CA Prop 65: Carcinogens", "Carcinogen",
    "Contains Batteries or Cells?", "Battery Type",
    "Safety Data Sheet (SDS) for flammable materials",
    "UN38.3 Test Summary or Safety Data Sheet (SDS) for products with batteries",
    "Safety Data Sheet (SDS) for aerosol products",
    "Safety Data Sheet (SDS) for other dangerous goods or hazardous materials",
    "Product Status",
]
_TIKTOK_ROW4 = [
    "Mandatory", "Optional", "Mandatory", "Mandatory", "Mandatory",
    "Optional", "Optional", "Optional", "Optional", "Optional", "Optional",
    "Optional", "Optional", "Optional", "Optional",
    "Conditionally mandatory ", "Conditionally mandatory ",
    "Optional", "Optional", "Optional", "Optional", "Optional",
    "Optional", "Optional", "Optional", "Optional",
    "Conditionally mandatory ", "Optional",
    "Mandatory", "Conditionally mandatory ", "Conditionally mandatory ",
    "Conditionally mandatory ", "Optional", "Mandatory", "Optional",
    "Mandatory", "Optional", "Conditionally mandatory ", "Optional",
    "Conditionally mandatory ", "Conditionally mandatory ",
    "Conditionally mandatory ", "Conditionally mandatory ",
    "Conditionally mandatory ", "Conditionally mandatory ",
    "Conditionally mandatory ", "Conditionally mandatory ",
    "Conditionally mandatory ", "Conditionally mandatory ",
    "Conditionally mandatory ", "Conditionally mandatory ",
    "Conditionally mandatory ", "Conditionally mandatory ",
    "Conditionally mandatory ", "Conditionally mandatory ",
    "Conditionally mandatory ", "Conditionally mandatory ",
    "Conditionally mandatory ", "Conditionally mandatory ",
    "Conditionally mandatory ", "Conditionally mandatory ",
    "Conditionally mandatory ", "Conditionally mandatory ",
    "Conditionally mandatory ", "Conditionally mandatory ",
    "Conditionally mandatory ", "Conditionally mandatory ",
    "Conditionally mandatory ", "Optional",
]

# ── Best Buy headers (267 columns) ────────────────────────────
_BB_HEADERS = [
    "Category Name", "Shop SKU", "Product Name", "GTIN", "Brand",
    "Description", "Front_Zoom", "Angle_Zoom", "Left_Zoom", "Back_Zoom",
    "Alt_View_Zoom_1", "Alt_View_Zoom_2", "Alt_View_Zoom_3", "Alt_View_Zoom_4",
    "Alt_View_Zoom_5", "Alt_View_Zoom_6", "Alt_View_Zoom_7", "Alt_View_Zoom_8",
    "Alt_View_Zoom_9", "Alt_View_Zoom_10",
    "Feature Bullets: 1: Title", "Feature Bullets: 1: Description",
    "Feature Bullets: 2: Title", "Feature Bullets: 2: Description",
    "Feature Bullets: 3: Title", "Feature Bullets: 3: Description",
    "Feature Bullets: 4: Title", "Feature Bullets: 4: Description",
    "Feature Bullets: 5: Title", "Feature Bullets: 5: Description",
    "Product Documents: 1: URL", "Product Documents: 1: Title",
    "Product Documents: 1: Description", "Product Documents: 1: Document Type",
    "Product Documents: 1: Language", "Product Documents: 1: Version",
    "Product Documents: 2: URL", "Product Documents: 2: Title",
    "Product Documents: 2: Description", "Product Documents: 2: Document Type",
    "Product Documents: 2: Language", "Product Documents: 2: Version",
    "Product Disclaimers: 1", "Model Number", "Model Number", "Color", "Color",
    "Box_Contents: 1", "Product Height", "Product Length", "Product Width",
    "Product Weight", "Recommended Minimum Age", "Material",
    "Warranty - Parts", "Warranty - Labor",
    "Is a Toy or Game Designed for Person Under 13 Years Old",
    "Image Sensor Type", "Image Sensor Size", "Total Pixels",
    "Image Resolution (Display)", "Zoom Capability", "Digital Zoom",
    "Optical Zoom", "In-Camera Image Stabilization", "Low Light/High Sensitivity",
    "Image File Format(s)", "Autofocus", "Eye Autofocus (Eye AF)",
    "Number of Autofocus Points (Up To)", "Shutter Speeds",
    "Minimum Aperture", "Maximum Aperture", "Minimum ISO", "Maximum ISO",
    "Maximum Expandable ISO", "Brightness Control", "Burst Mode", "Burst Rate",
    "Self Timer", "Self-Timer Delay",
    "Trade Item Hierarchy (Configuration of the Supply Chain Packaging): Each Item: Dimensions: Length/Depth",
    "Shooting Modes",
    "Trade Item Hierarchy (Configuration of the Supply Chain Packaging): Each Item: Dimensions: Width",
    "Panorama Mode",
    "Trade Item Hierarchy (Configuration of the Supply Chain Packaging): Each Item: Dimensions: Height",
    "Scene Modes",
    "Trade Item Hierarchy (Configuration of the Supply Chain Packaging): Each Item: Weight: Amount",
    "White Balance Modes", "Last Return Date", "Effective Pixels",
    "California Proposition 65 Warning: Message",
    "Lens Mount Compatibility", "Lens Type", "Minimum Focal Length",
    "Maximum Focal Length", "Craft Project Type", "In-Lens Image Stabilization",
    "League/Association", "Lens Series", "Brand/Character", "Lens Model Number",
    "CPSC Choking Hazard", "Lens 2 Type",
    "Country: Primary: Country of Origin", "Lens 2 Minimum Aperture",
    "Trade Item Hierarchy (Configuration of the Supply Chain Packaging): Each Item: Dimensions: Unit of Measure",
    "Lens 2 Maximum Aperture",
    "Trade Item Hierarchy (Configuration of the Supply Chain Packaging): Each Item: Weight: Unit of Measure",
    "Lens 2 Minimum Focal Length", "California Proposition 65 Warning: Type",
    "Lens 2 Maximum Focal Length", "Contains Embedded Battery",
    "Lens 2 Series", "Lens 2 Model Number", "Lens 3 Type",
    "Lens 3 Minimum Aperture", "Lens 3 Maximum Aperture",
    "Lens 3 Minimum Focal Length", "Lens 3 Maximum Focal Length",
    "Lens 3 Series", "Lens 3 Model Number", "Lens 4 Type",
    "Lens 4 Minimum Aperture", "Lens 4 Maximum Aperture",
    "Lens 4 Minimum Focal Length", "Lens 4 Maximum Focal Length",
    "Lens 4 Series", "Lens 4 Model Number", "Flash Modes", "Guide Number",
    "Integrated Flash", "External Flash Mount", "Display Type", "Screen Size",
    "Aspect Ratio", "Screen Resolution", "Viewfinder Type",
    "Viewfinder Magnification", "Touch Screen", "Varying Angle Screen",
    "Standardized Video Resolution", "Video Resolution(s) and Frame Rate(s)",
    "Maximum Video Frame Rate", "Slow Motion Recording",
    "HD Movie Mode Focus", "Number of Memory Card Slots",
    "Tripod/Monopod Mountable", "Socket Size", "Memory Card Compatibility",
    "Remote Compatible", "Face Detection", "Smile Mode", "Integrated GPS",
    "Records Audio", "Wireless Connectivity", "Output(s)", "Microphone Input",
    "Headphone Jack", "PictBridge Enabled", "Rechargeable Battery",
    "Battery Life (up to)", "AC Adapter Compatible", "Removable Battery",
    "Battery Model", "Number of Batteries Required",
    "Number of Images Per Charge", "Protective Qualities",
    "Maximum Depth of Water Resistance", "Maximum Height of Shock Resistance",
    "Number of Lenses Included", "Camera Bag/Case Included",
    "Batteries Included", "AC Adapter Included", "Memory Card Included",
    "Product Height", "Product Width", "Product Depth", "Product Weight",
    "Filter Diameter", "Camera Body Weight", "Charging Interface(s)",
    "Warranty - Parts", "Warranty - Labor", "Product Set",
    "Camera Model Family", "Digital Camera Type", "CPSC Choking Hazard",
    "Is a Toy or Game Designed for Person Under 13 Years Old",
    "Country: Primary: Country of Origin",
    "Trade Item Hierarchy (Configuration of the Supply Chain Packaging): Each Item: Dimensions: Length/Depth",
    "Trade Item Hierarchy (Configuration of the Supply Chain Packaging): Each Item: Dimensions: Width",
    "Trade Item Hierarchy (Configuration of the Supply Chain Packaging): Each Item: Dimensions: Height",
    "Trade Item Hierarchy (Configuration of the Supply Chain Packaging): Each Item: Dimensions: Unit of Measure",
    "Trade Item Hierarchy (Configuration of the Supply Chain Packaging): Each Item: Weight: Amount",
    "Trade Item Hierarchy (Configuration of the Supply Chain Packaging): Each Item: Weight: Unit of Measure",
    "Last Return Date", "California Proposition 65 Warning: Type",
    "California Proposition 65 Warning: Message",
    "Contains intentionally added PFAS subject to state sales restrictions",
    "Box_Contents: 1", "HDR Mode", "Contains Embedded Battery",
    "Model Number", "Color", "Material", "BPA-Free", "Phthalate Free",
    "JPMA Certified", "Product Height", "Product Width", "Product Length",
    "Product Weight", "Assembly Required", "Care And Cleaning Instructions",
    "Batteries Included", "Battery Chemistry", "Recommended Minimum Age",
    "Recommended Maximum Age", "Battery Size", "Number of Batteries Required",
    "On/Off Switch", "Fill Material", "Warranty - Parts", "Warranty - Labor",
    "Function(s)", "Battery Powered", "Plush Toy Type", "Replaceable Battery",
    "Brand/Character", "CPSC Choking Hazard",
    "Is a Toy or Game Designed for Person Under 13 Years Old",
    "Consumer Lifestage", "Country: Primary: Country of Origin",
    "Trade Item Hierarchy (Configuration of the Supply Chain Packaging): Each Item: Dimensions: Length/Depth",
    "Trade Item Hierarchy (Configuration of the Supply Chain Packaging): Each Item: Dimensions: Width",
    "Trade Item Hierarchy (Configuration of the Supply Chain Packaging): Each Item: Dimensions: Height",
    "Trade Item Hierarchy (Configuration of the Supply Chain Packaging): Each Item: Dimensions: Unit of Measure",
    "Trade Item Hierarchy (Configuration of the Supply Chain Packaging): Each Item: Weight: Amount",
    "Trade Item Hierarchy (Configuration of the Supply Chain Packaging): Each Item: Weight: Unit of Measure",
    "Last Return Date", "California Proposition 65 Warning: Type",
    "California Proposition 65 Warning: Message",
    "Contains intentionally added PFAS subject to state sales restrictions",
    "BPA Free", "Box_Contents: 1", "League/Association",
    "Offer SKU", "Product ID", "Product ID Type",
    "Offer Description", "Offer Internal Description", "Offer Price",
    "Offer Price Additional Info", "Offer Quantity", "Minimum Quantity Alert",
    "Offer State", "Availability Start Date", "Availability End Date",
    "Logistic Class", "Favorite Rank", "Discount price",
    "Discount Start Date", "Discount End Date",
    "Lead Time to Ship (in days)", "Update/Delete", "Product Tax Code",
]

# ── Toys R Us / Logicbroker headers (38 columns) ──────────────
_TRU_HEADERS = [
    "error", "master_product_id", "product_id", "display_name",
    "ean", "upc", "short_Description", "long_Description", "brand",
    "manufacturer_name", "manufacturer_sku", "min_order_quantity",
    "step_quantity", "imageUrl_0", "imageUrl_1", "imageUrl_2",
    "imageUrl_3", "imageUrl_4", "AMZManufacturerSKU", "esrb",
    "finish_color", "mpaarating", "platform", "release_date", "wheelSize",
    "shipmentMessage", "cost", "msrp", "regularPrice", "salePrice",
    "vendorId", "color", "size", "AMZManufacturerName",
    "startDate", "endDate", "recommendedAge", "learningSkill",
]

# ── Target Plus headers (13 columns) ──────────────────────────
_TARGET_HEADERS = [
    "SKU", "Listing ID", "Barcode", "Category", "Import Description",
    "Prop 65", "Tax",
    "Bullet Features", "Bullet Features", "Bullet Features",
    "Bullet Features", "Bullet Features",
    "Shipping Exclusion",
]

# ── eBay File Exchange headers (47 columns) ───────────────────
_EBAY_HEADERS = [
    "Line Number", "Action", "Status", "ErrorCode", "ErrorMessage",
    "WarningCode", "WarningMessage", "Code", "Message", "ItemID",
    "ReferenceID", "ApplicationData", "StartTime", "EndTime",
    "AuctionLengthFee", "BoldFee", "BorderFee", "BuyItNowFee",
    "CategoryFeaturedFee", "CurrencyID", "FeaturedFee",
    "FeaturedGalleryFee", "FixedPriceDurationFee", "GalleryFee",
    "GiftIconFee", "HighlightFee", "InsertionFee",
    "InternationalInsertionFee", "ListingDesignerFee", "ListingFee",
    "PhotoDisplayFee", "PhotoFee", "ProPackBundleFee", "ReserveFee",
    "SchedulingFee", "SubtitleFee", "CustomLabel", "PrivateNotes",
    "BasicUpgradePackBundleFee", "ValuePackBundleFee",
    "ProPackPlusBundleFee", "SellerInventoryID",
    "CrossBorderTradeNorthAmericaFee", "CrossBorderTradeGBFee",
    "RefundFromSeller", "TotalRefundToBuyer", "CorrelationID",
]


def generate_walmart_xlsx(products, supplier_id, output_path):
    """Walmart Product Content And Site Exp upload file."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Product Content And Site Exp"

    for _ in range(3):
        ws.append([""] * 228)
    ws.append(coerce_row(_WM_HEADERS))

    for p in products:
        row = [""] * 228
        row[3]   = "UPC"
        row[4]   = p["upc"]
        row[6]   = p["ds_number"]
        row[7]   = supplier_id
        row[8]   = safe_str(p["product_name"])
        row[9]   = safe_str(p["brand"])
        row[11]  = p["msrp"]
        row[12]  = p["drop_ship_cost"]
        row[24]  = safe_str(p["country_of_origin"])
        row[28]  = safe_str(p["battery_type_qty"])
        row[29]  = p["pkg_length"]
        row[30]  = p["pkg_width"]
        row[31]  = p["pkg_height"]
        row[32]  = p["pkg_weight"]
        row[37]  = safe_str(p["description"])
        row[64]  = safe_str(p["primary_color"])
        row[72]  = safe_str(p["bullet1"])
        row[73]  = safe_str(p["bullet2"])
        row[74]  = safe_str(p["bullet3"])
        row[75]  = safe_str(p["bullet4"])
        row[76]  = safe_str(p["bullet5"])
        row[81]  = safe_str(p["image_url_1"])
        row[82]  = safe_str(p["material"])
        row[109] = p["num_pieces"]
        row[179] = safe_str(p["brand"])
        row[180] = safe_str(p["vendor_item_number"])
        row[203] = safe_str(p["product_name"])
        ws.append(coerce_row(row))

    _safe_save_wb(wb, output_path)
    print(f"  ✓ Walmart          → {os.path.basename(output_path)}  ({len(products)} item(s))")


def generate_tiktok_xlsx(products, tiktok_category, output_path):
    """TikTok Shop product listing template."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Template"

    ws.append(coerce_row(_TIKTOK_ROW1))
    ws.append(coerce_row(_TIKTOK_ROW2))
    ws.append(coerce_row(_TIKTOK_ROW3))
    ws.append(coerce_row(_TIKTOK_ROW4))

    for p in products:
        row = [""] * 69
        row[0]  = tiktok_category
        row[1]  = safe_str(p["brand"])
        row[2]  = safe_str(p["product_name"])
        row[3]  = safe_str(p["description"])
        row[4]  = safe_str(p["image_url_1"])
        row[5]  = safe_str(p["image_url_2"])
        row[6]  = safe_str(p["image_url_3"])
        row[7]  = safe_str(p["image_url_4"])
        row[8]  = safe_str(p["image_url_5"])
        row[13] = safe_str(p["id_type"])
        row[14] = p["upc"]
        row[28] = p["pkg_weight"]
        row[29] = p["pkg_length"]
        row[30] = p["pkg_width"]
        row[31] = p["pkg_height"]
        row[33] = p["msrp"]
        row[36] = p["ds_number"]
        row[41] = safe_str(p["country_of_origin"])
        row[42] = safe_str(p["min_age"])
        row[43] = safe_str(p["bullet1"])
        row[44] = safe_str(p["brand"])
        row[62] = safe_str(p["batteries_required"])
        row[63] = safe_str(p["battery_type_qty"])
        ws.append(coerce_row(row))

    _safe_save_wb(wb, output_path)
    print(f"  ✓ TikTok Shop      → {os.path.basename(output_path)}  ({len(products)} item(s))")


def generate_bestbuy_xlsx(products, bb_category, output_path):
    """Best Buy product + offer upload file."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"

    ws.append(coerce_row(_BB_HEADERS))

    for p in products:
        row = [""] * 267

        row[0]  = bb_category
        row[1]  = p["ds_number"]
        row[2]  = safe_str(p["product_name"])
        row[3]  = p["upc"]
        row[4]  = safe_str(p["brand"])
        row[5]  = safe_str(p["description"])
        row[6]  = safe_str(p["image_url_1"])
        row[7]  = safe_str(p["image_url_2"])
        row[8]  = safe_str(p["image_url_3"])
        row[9]  = safe_str(p["image_url_4"])
        row[10] = safe_str(p["image_url_5"])
        row[20] = ""
        row[21] = safe_str(p["bullet1"])
        row[22] = ""
        row[23] = safe_str(p["bullet2"])
        row[24] = ""
        row[25] = safe_str(p["bullet3"])
        row[26] = ""
        row[27] = safe_str(p["bullet4"])
        row[28] = ""
        row[29] = safe_str(p["bullet5"])
        row[43] = safe_str(p["vendor_item_number"])
        row[48] = p["item_height"]
        row[49] = p["item_length"]
        row[50] = p["item_width"]
        row[51] = p["item_weight"]
        row[52] = safe_str(p["min_age"])
        row[53] = safe_str(p["material"])
        row[102] = safe_str(p["choking_hazard"])
        row[104] = safe_str(p["country_of_origin"])
        row[207] = safe_str(p["vendor_item_number"])
        row[209] = p["item_height"]
        row[210] = p["item_width"]
        row[211] = p["item_length"]
        row[212] = p["item_weight"]
        row[213] = safe_str(p["assembly_required"])
        row[215] = safe_str(p["batteries_included"])
        row[216] = safe_str(p["battery_cell_comp"])
        row[217] = safe_str(p["min_age"])
        row[218] = safe_str(p["max_age"])
        row[230] = safe_str(p["choking_hazard"])
        row[233] = safe_str(p["country_of_origin"])
        row[247] = p["ds_number"]
        row[248] = p["upc"]
        row[249] = "UPC"
        row[252] = p["msrp"]
        row[254] = 1
        ws.append(coerce_row(row))

    _safe_save_wb(wb, output_path)
    print(f"  ✓ Best Buy         → {os.path.basename(output_path)}  ({len(products)} item(s))")


def generate_toysrus_xlsx(products, output_path):
    """Toys R Us / Logicbroker product upload file."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Product"

    ws.append(coerce_row(_TRU_HEADERS))

    for p in products:
        row = [""] * 38
        row[2]  = p["ds_number"]
        row[3]  = safe_str(p["product_name"])
        row[5]  = p["upc"]
        row[6]  = safe_str(p["bullet1"])
        row[7]  = safe_str(p["description"])
        row[8]  = safe_str(p["brand"])
        row[9]  = safe_str(p["brand"])
        row[10] = safe_str(p["vendor_item_number"])
        row[11] = "1"
        row[12] = "1"
        row[13] = safe_str(p["image_url_1"])
        row[14] = safe_str(p["image_url_2"])
        row[15] = safe_str(p["image_url_3"])
        row[16] = safe_str(p["image_url_4"])
        row[17] = safe_str(p["image_url_5"])
        row[26] = p["drop_ship_cost"]
        row[27] = p["msrp"]
        row[28] = p["msrp"]
        row[36] = safe_str(p["min_age"])
        ws.append(coerce_row(row))

    _safe_save_wb(wb, output_path)
    print(f"  ✓ Toys R Us        → {os.path.basename(output_path)}  ({len(products)} item(s))")


def generate_target_csv(products, target_category, output_path):
    """Target Plus listing-attribute CSV."""
    tmp_fd, tmp_path = tempfile.mkstemp(suffix=".csv")
    try:
        with os.fdopen(tmp_fd, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(_TARGET_HEADERS)
            for p in products:
                writer.writerow([
                    p["ds_number"],
                    "",
                    p["upc"],
                    target_category,
                    safe_str(p["description"]),
                    "",
                    "",
                    safe_str(p["bullet1"]),
                    safe_str(p["bullet2"]),
                    safe_str(p["bullet3"]),
                    safe_str(p["bullet4"]),
                    safe_str(p["bullet5"]),
                    "",
                ])
        shutil.copy2(tmp_path, output_path)
    finally:
        if os.path.exists(tmp_path):
            os.remove(tmp_path)

    print(f"  ✓ Target Plus      → {os.path.basename(output_path)}  ({len(products)} item(s))")


def generate_ebay_csv(products, output_path):
    """eBay File Exchange listing CSV."""
    tmp_fd, tmp_path = tempfile.mkstemp(suffix=".csv")
    try:
        with os.fdopen(tmp_fd, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(_EBAY_HEADERS)
            for i, p in enumerate(products, start=1):
                row = [""] * 47
                row[0]  = i
                row[1]  = "Add"
                row[36] = p["ds_number"]
                writer.writerow(row)
        shutil.copy2(tmp_path, output_path)
    finally:
        if os.path.exists(tmp_path):
            os.remove(tmp_path)

    print(f"  ✓ eBay             → {os.path.basename(output_path)}  ({len(products)} item(s))")


_SHOPIFY_HEADERS = [
    "Handle", "Title", "Body (HTML)", "Vendor", "Product Category",
    "Type", "Tags", "Published", "Option1 Name", "Option1 Value",
    "Option1 Linked To", "Option2 Name", "Option2 Value", "Option2 Linked To",
    "Option3 Name", "Option3 Value", "Option3 Linked To", "Variant SKU",
    "Variant Grams", "Variant Inventory Tracker", "Variant Inventory Policy",
    "Variant Fulfillment Service", "Variant Price", "Variant Compare At Price",
    "Variant Requires Shipping", "Variant Taxable",
    "Unit Price Total Measure", "Unit Price Total Measure Unit",
    "Unit Price Base Measure", "Unit Price Base Measure Unit",
    "Variant Barcode", "Image Src", "Image Position", "Image Alt Text",
    "Gift Card", "SEO Title", "SEO Description",
    "bullet_1 (product.metafields.custom.bullet_1)",
    "bullet_2 (product.metafields.custom.bullet_2)",
    "bullet_3 (product.metafields.custom.bullet_3)",
    "Details & Specs (product.metafields.custom.details_and_specs)",
    "ds_sku (product.metafields.custom.ds_sku)",
    "height_in (product.metafields.custom.height_in)",
    "length_in (product.metafields.custom.length_in)",
    "license (product.metafields.custom.license)",
    "Shipping & Returns (product.metafields.custom.shipping_and_returns)",
    "width_in (product.metafields.custom.width_in)",
    "Variant Image", "Variant Weight Unit", "Variant Tax Code",
    "Cost per item", "Status",
]


def generate_shopify_xlsx(products, shopify_type, output_path):
    """
    Shopify product import file.
    52 columns matching the griomax upload template.
    One row per product. Weight converted to grams; dims in inches.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(_SHOPIFY_HEADERS)

    for p in products:
        row = [""] * 52

        # ── Identity & content ─────────────────────────────────
        row[0]  = shopify_handle(p["product_name"])                     # Handle
        row[1]  = safe_str(p["product_name"])                           # Title
        row[2]  = f"<p>{safe_str(p['description'])}</p>"                # Body (HTML)
        row[3]  = safe_str(p["brand"])                                   # Vendor
        row[4]  = ""                                                     # Product Category (user fills)
        row[5]  = shopify_type                                           # Type
        row[6]  = safe_str(p["keywords"])                               # Tags
        row[7]  = True                                                   # Published

        # ── Options (single-variant) ────────────────────────────
        row[8]  = "Title"                                                # Option1 Name
        row[9]  = "Default Title"                                        # Option1 Value

        # ── Variant fields ──────────────────────────────────────
        row[17] = str(p["ds_number"]) if p["ds_number"] else ""          # Variant SKU
        row[18] = to_grams(p["pkg_weight"], p["pkg_weight_unit"])        # Variant Grams
        row[19] = "shopify"                                              # Variant Inventory Tracker
        row[20] = "deny"                                                 # Variant Inventory Policy
        row[21] = "manual"                                               # Variant Fulfillment Service
        row[22] = p["msrp"]                                              # Variant Price
        row[24] = True                                                   # Variant Requires Shipping
        row[25] = True                                                   # Variant Taxable
        row[30] = safe_str(p["upc"])                                     # Variant Barcode

        # ── Media ───────────────────────────────────────────────
        row[31] = safe_str(p["image_url_1"])                             # Image Src

        # ── Other fixed fields ──────────────────────────────────
        row[34] = False                                                  # Gift Card
        row[35] = safe_str(p["product_name"])                           # SEO Title
        row[36] = safe_str(p["description"], max_len=320)               # SEO Description

        # ── Custom metafields ───────────────────────────────────
        row[37] = safe_str(p["bullet1"])                                 # bullet_1
        row[38] = safe_str(p["bullet2"])                                 # bullet_2
        row[39] = safe_str(p["bullet3"])                                 # bullet_3
        # [40] Details & Specs — left blank
        try:
            row[41] = int(p["ds_number"])                               # ds_sku (integer)
        except (TypeError, ValueError):
            row[41] = p["ds_number"]
        row[42] = dim_to_inches(p["pkg_height"], p["pkg_height_unit"])  # height_in
        row[43] = dim_to_inches(p["pkg_length"], p["pkg_length_unit"])  # length_in
        row[44] = safe_str(p["brand"])                                   # license
        # [45] Shipping & Returns — left blank
        row[46] = dim_to_inches(p["pkg_width"],  p["pkg_width_unit"])   # width_in

        # ── Variant weight / cost / status ──────────────────────
        row[48] = "g"                                                    # Variant Weight Unit
        row[50] = p["drop_ship_cost"]                                    # Cost per item
        row[51] = "active"                                               # Status

        ws.append(coerce_row(row))

    _safe_save_wb(wb, output_path)
    print(f"  ✓ Shopify          → {os.path.basename(output_path)}  ({len(products)} item(s))")


def generate_channel_files(products, supplier_id, base, output_dir):
    """Prompt for channel-specific inputs and generate all 7 channel setup files."""
    print()
    print("  Channel setup — enter category/classification for each portal:")
    tiktok_cat   = input("  TikTok Shop category: ").strip()
    bb_cat       = input("  Best Buy category name: ").strip()
    target_cat   = input("  Target Plus category: ").strip()
    shopify_type = input("  Shopify product type (e.g. Toy, Action Figure): ").strip()

    print(f"\n  Generating channel setup files...")
    generate_walmart_xlsx(
        products, supplier_id,
        os.path.join(output_dir, f"{base}_walmart.xlsx"),
    )
    generate_tiktok_xlsx(
        products, tiktok_cat,
        os.path.join(output_dir, f"{base}_tiktok.xlsx"),
    )
    generate_bestbuy_xlsx(
        products, bb_cat,
        os.path.join(output_dir, f"{base}_bestbuy.xlsx"),
    )
    generate_toysrus_xlsx(
        products,
        os.path.join(output_dir, f"{base}_toysrus.xlsx"),
    )
    generate_target_csv(
        products, target_cat,
        os.path.join(output_dir, f"{base}_target.csv"),
    )
    generate_ebay_csv(
        products,
        os.path.join(output_dir, f"{base}_ebay.csv"),
    )
    generate_shopify_xlsx(
        products, shopify_type,
        os.path.join(output_dir, f"{base}_shopify.xlsx"),
    )


# ══════════════════════════════════════════════════════════════
#  MAIN
# ══════════════════════════════════════════════════════════════

def main():
    print("=" * 60)
    print("  DS WMS Upload File Generator  (2026 Template)")
    print("=" * 60)

    if len(sys.argv) > 1:
        template_path = sys.argv[1].strip('"').strip("'")
    else:
        template_path = input("\nDrag and drop the New Item Setup Sheet here\n(or type the full path): ").strip().strip('"').strip("'")

    if not os.path.isfile(template_path):
        print(f"\nERROR: File not found → {template_path}")
        input("Press Enter to exit.")
        sys.exit(1)

    # ── Collect the non-supplier prompts first ─────────────────
    category = input("Product Category (e.g. Toys): ").strip()

    lead_time_raw = input(f"Lead time in days [{DEFAULT_LEAD_TIME}]: ").strip()
    lead_time = int(lead_time_raw) if lead_time_raw else DEFAULT_LEAD_TIME

    moq_raw = input(f"MOQ (min order quantity) [{DEFAULT_MOQ}]: ").strip()
    moq = int(moq_raw) if moq_raw else DEFAULT_MOQ

    print()
    new_ds_raw = input(
        "First NEW DS number (Apprise New Product upload only — items below\n"
        "  this DS# are already in the system and will be skipped).\n"
        "  Enter 0 or leave blank to include all items: "
    ).strip()
    try:
        new_ds_start = int(new_ds_raw) if new_ds_raw else 0
    except ValueError:
        new_ds_start = 0

    # ── Read template (also returns supplier ID from B3) ───────
    try:
        products, sheet_supplier_id = read_template(template_path)
    except Exception as e:
        print(f"\nERROR reading template: {e}")
        input("Press Enter to exit.")
        sys.exit(1)

    if not products:
        print("\nNo products found.")
        print("Make sure DS Numbers are entered in Column A of the 'DS Only' sheet.")
        input("Press Enter to exit.")
        sys.exit(1)

    # ── Resolve Supplier ID ────────────────────────────────────
    print()
    if sheet_supplier_id is not None:
        print(f"  Apprise Supplier Number read from sheet: {sheet_supplier_id}")
        supplier_id = sheet_supplier_id
    else:
        print("  ⚠ Apprise Supplier Number in cell B3 is blank or set to 'xxxx'.")
        supplier_id_raw = input("  Enter Supplier ID now (e.g. 1514): ").strip()
        try:
            supplier_id = int(supplier_id_raw)
        except ValueError:
            supplier_id = supplier_id_raw

    print(f"\n{len(products)} product(s) found. Generating WMS upload files...")

    output_dir = os.path.dirname(os.path.abspath(template_path))
    base       = os.path.splitext(os.path.basename(template_path))[0]
    np_path    = os.path.join(output_dir, f"{base}_new_product.xlsx")

    try:
        generate_new_product(
            products, supplier_id, category,
            np_path,
            min_ds_number=new_ds_start,
        )
        generate_supplier(
            products, supplier_id, lead_time, moq,
            os.path.join(output_dir, f"{base}_supplier.xlsx")
        )
        generate_uom(
            products,
            os.path.join(output_dir, f"{base}_uom.xlsx")
        )
        generate_purchase_cost(
            products, supplier_id,
            os.path.join(output_dir, f"{base}_purchase_cost.xlsx")
        )
        generate_goflow_csv(
            products,
            os.path.join(output_dir, f"{base}_goflow.csv")
        )
    except Exception as e:
        print(f"\nERROR generating files: {e}")
        input("Press Enter to exit.")
        sys.exit(1)

    validate_new_product(products, np_path, min_ds_number=new_ds_start)

    print(f"\n{'=' * 60}")
    print(f"  Done! 5 WMS files saved to:")
    print(f"  {output_dir}")
    print(f"{'=' * 60}")

    print()
    gen_channels = input("Generate channel setup files? (Y/N) [N]: ").strip().lower()
    if gen_channels == "y":
        try:
            generate_channel_files(products, supplier_id, base, output_dir)
        except Exception as e:
            print(f"\nERROR generating channel files: {e}")
            input("Press Enter to exit.")
            sys.exit(1)

        print(f"\n{'=' * 60}")
        print(f"  Done! Channel files saved to:")
        print(f"  {output_dir}")
        print(f"{'=' * 60}")

    input("\nPress Enter to exit.")


if __name__ == "__main__":
    main()
