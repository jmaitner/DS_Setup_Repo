#!/usr/bin/env python3
"""
ingest_status.py  —  Update the status layer in bulk from a channel's export / error report.

You hand over a spreadsheet a channel gave you (a listing report, an error/quality report,
a "what's live" export). This matches each row to a product by DS# (or by SKU = vendor item
number, or by the channel listing id), and updates that product's status for ONE channel.

Channel report formats vary, so columns are mapped with flags. Typical use: Claude inspects
the file's headers, then runs this with the right column names.

Examples:
  # Mark everything in a Walmart "live items" export as live on walmart_1p, keyed by SKU
  python scripts/ingest_status.py "walmart_live.xlsx" --channel walmart_1p \\
      --match-type sku --match-col "Seller SKU" --set-state live --id-col "Item ID"

  # Load an Amazon error report, key by DS#, capture the error text + case number
  python scripts/ingest_status.py "amzn_errors.xlsx" --channel amazon \\
      --match-col "DS Number" --set-state error --issue-col "Error Message" --case-col "Case ID"

Non-destructive-ish: updates only the named channel's fields; reports unmatched rows.
Use --dry-run to preview.
"""

import argparse
import glob
import json
import os
import sys
from datetime import date

for _stream in (sys.stdout, sys.stderr):
    try:
        _stream.reconfigure(encoding="utf-8")
    except (AttributeError, ValueError):
        pass

try:
    from openpyxl import load_workbook
except ImportError:
    sys.exit("ERROR: openpyxl not installed. Run: pip install -r requirements.txt")

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from status_lib import (CHANNELS, CHANNEL_STATES, PRODUCTS_DIR,
                        new_status_doc, load_status, save_status)


def build_lookups():
    """Return (ds_set, sku_to_ds). SKU = vendor_item_number."""
    ds_set, sku_to_ds = set(), {}
    for path in glob.glob(os.path.join(PRODUCTS_DIR, "**", "*.json"), recursive=True):
        with open(path, encoding="utf-8") as f:
            p = json.load(f)
        ds = str(p.get("ds_number") or "").strip()
        if not ds:
            continue
        ds_set.add(ds)
        sku = p.get("vendor_item_number")
        if sku:
            sku_to_ds[str(sku).strip().lower()] = ds
    return ds_set, sku_to_ds


def find_header_row(ws, want, max_scan=8):
    want = want.strip().lower()
    for r in range(1, min(max_scan, ws.max_row) + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str) and v.strip().lower() == want:
                return r
    return None


def col_index(ws, hdr_row, name):
    if not name:
        return None
    name = name.strip().lower()
    for c in range(1, ws.max_column + 1):
        v = ws.cell(hdr_row, c).value
        if isinstance(v, str) and v.strip().lower() == name:
            return c
    return None


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("xlsx")
    ap.add_argument("--channel", required=True, choices=list(CHANNELS))
    ap.add_argument("--match-type", choices=["ds", "sku", "id"], default="ds")
    ap.add_argument("--match-col", help="header of the key column (auto-detected if omitted)")
    ap.add_argument("--set-state", choices=CHANNEL_STATES, help="state to set on matched rows")
    ap.add_argument("--state-col", help="header whose value becomes the channel state")
    ap.add_argument("--issue-col", help="header whose value becomes the issue text")
    ap.add_argument("--case-col", help="header whose value becomes the case number")
    ap.add_argument("--id-col", help="header whose value becomes the channel listing id")
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()

    path = args.xlsx.strip().strip('"').strip("'")
    if not os.path.isfile(path):
        sys.exit(f"ERROR: file not found -> {path}")

    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]

    default_key = {"ds": "DS Number", "sku": "SKU", "id": "Item ID"}[args.match_type]
    key_name = args.match_col or default_key
    hdr_row = find_header_row(ws, key_name)
    if not hdr_row:
        sys.exit(f"ERROR: header '{key_name}' not found in first sheet. "
                 f"Pass --match-col with the exact key column header.")

    key_col = col_index(ws, hdr_row, key_name)
    cols = {f: col_index(ws, hdr_row, getattr(args, f"{f}_col"))
            for f in ("state", "issue", "case", "id")}

    ds_set, sku_to_ds = build_lookups()
    today = date.today().isoformat()
    matched, unmatched, updated = [], [], []

    for r in range(hdr_row + 1, ws.max_row + 1):
        raw = ws.cell(r, key_col).value
        if raw is None or str(raw).strip() == "":
            continue
        key = str(raw).strip()
        if key.endswith(".0"):
            key = key[:-2]

        if args.match_type == "sku":
            ds = sku_to_ds.get(key.lower())
        else:  # ds (or id treated as ds fallback)
            ds = key if key in ds_set else None

        if not ds:
            unmatched.append(key)
            continue
        matched.append(ds)

        def cell(field):
            c = cols[field]
            if not c:
                return None
            v = ws.cell(r, c).value
            return str(v).strip() if v not in (None, "") else None

        doc = load_status(ds) or new_status_doc(ds, today)
        ch = doc["channels"].setdefault(args.channel, {})
        state = (cell("state") or args.set_state)
        if state:
            ch["state"] = state if state in CHANNEL_STATES else ch.get("state", "error")
        if cell("id"):    ch["id"] = cell("id")
        if cell("case"):  ch["case_number"] = cell("case")
        if cell("issue"): ch["issue"] = cell("issue")
        ch["updated"] = today
        doc["updated"] = today

        if args.dry_run:
            updated.append(ds)
        else:
            save_status(doc)
            updated.append(ds)

    print(f"\nChannel: {args.channel}   matched: {len(set(matched))}   "
          f"updated: {len(set(updated))}   unmatched rows: {len(unmatched)}")
    if unmatched:
        print("  Unmatched keys (no product): "
              + ", ".join(unmatched[:20]) + (" ..." if len(unmatched) > 20 else ""))
    if not args.dry_run:
        print("Next: python scripts/validate.py && python scripts/build_index.py")


if __name__ == "__main__":
    main()
