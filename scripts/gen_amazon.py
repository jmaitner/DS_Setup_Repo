#!/usr/bin/env python3
"""
gen_amazon.py  —  Generate an Amazon flat-file (inventory loader) from catalog products.

Two modes:
  • Generic (default): writes an .xlsx using Amazon's standard flat-file field names
    (item_sku, item_name, brand_name, external_product_id, bullet_point1..5, ...).
    Immediately usable as a data file / source to paste into a template.
  • Template (--template "Amazon Toys template.xlsx"): fills Amazon's OWN downloaded
    category template in place — finds the field-name header row (the one containing
    'item_sku') on the 'Template' sheet, maps our data into the matching columns, and
    appends data rows below it, preserving Amazon's metadata header rows. This produces
    a file Amazon will accept for that category.

Skips discontinued items. Adds a human-readable category recommendation.

Examples:
  python scripts/gen_amazon.py --vendor "Headstart Inc" --out "C:\\...\\Headstart Amazon.xlsx"
  python scripts/gen_amazon.py --vendor "Headstart Inc" --template "Toys.xlsm" --out out.xlsx
"""

import argparse
import glob
import json
import os
import sys

for _stream in (sys.stdout, sys.stderr):
    try:
        _stream.reconfigure(encoding="utf-8")
    except (AttributeError, ValueError):
        pass

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("ERROR: openpyxl not installed. Run: pip install -r requirements.txt")

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from ds_schema import brand_slug

REPO_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
PRODUCTS_DIR = os.path.join(REPO_ROOT, "products")
STATUS_DIR = os.path.join(REPO_ROOT, "status")

# Standard Amazon flat-file field names we can populate from the catalog.
AMAZON_FIELDS = [
    "feed_product_type", "item_sku", "brand_name", "item_name", "manufacturer",
    "part_number", "external_product_id", "external_product_id_type", "standard_price",
    "quantity", "main_image_url", "other_image_url1", "other_image_url2",
    "other_image_url3", "other_image_url4", "bullet_point1", "bullet_point2",
    "bullet_point3", "bullet_point4", "bullet_point5", "product_description",
    "recommended_browse_nodes", "color_name", "material_type", "age_range_description",
    "country_of_origin", "are_batteries_required", "item_weight",
    "item_weight_unit_of_measure", "item_length", "item_width", "item_height",
    "item_length_unit_of_measure", "condition_type",
]


def amazon_category(name):
    n = (name or "").lower()
    if "accessor" in n or "necklace" in n:
        return "Toys & Games > Dolls & Accessories > Doll Playsets & Accessories"
    if "doll" in n:
        return "Toys & Games > Dolls & Accessories > Dolls"
    if "plush" in n:
        return "Toys & Games > Stuffed Animals & Plush Toys"
    if "walker" in n:
        return "Baby Products > Activity & Entertainment > Walkers"
    if any(k in n for k in ("sorter", "shape o", "stack")):
        return "Toys & Games > Baby & Toddler Toys > Sorting, Stacking & Plugging Toys"
    if "vehicle" in n or "push n play" in n:
        return "Toys & Games > Play Vehicles"
    if "learn" in n or "bundle" in n:
        return "Toys & Games > Baby & Toddler Toys > Early Development & Activity Toys"
    return "Toys & Games > Baby & Toddler Toys"


def load_lifecycles():
    out = {}
    for p in glob.glob(os.path.join(STATUS_DIR, "*.json")):
        d = json.load(open(p, encoding="utf-8"))
        out[d["ds_number"]] = d.get("lifecycle")
    return out


def row_for(prod):
    """Map a product JSON to a dict of Amazon field -> value."""
    b = ((prod.get("content") or {}).get("bullets") or []) + [""] * 5
    imgs = ((prod.get("images") or {}).get("urls") or []) + [""] * 5
    a = prod.get("attributes") or {}
    pr = prod.get("pricing") or {}
    src = prod.get("sourcing") or {}
    it = (prod.get("dimensions") or {}).get("item") or {}
    comp = prod.get("compliance") or {}
    return {
        "feed_product_type": "toys",
        "item_sku": prod.get("ds_number"),
        "brand_name": prod.get("brand"),
        "item_name": prod.get("product_name"),
        "manufacturer": prod.get("brand") or prod.get("vendor"),
        "part_number": prod.get("vendor_item_number"),
        "external_product_id": (prod.get("identity") or {}).get("upc"),
        "external_product_id_type": "UPC",
        "standard_price": pr.get("msrp"),
        "quantity": "",
        "main_image_url": imgs[0],
        "other_image_url1": imgs[1], "other_image_url2": imgs[2],
        "other_image_url3": imgs[3], "other_image_url4": imgs[4],
        "bullet_point1": b[0], "bullet_point2": b[1], "bullet_point3": b[2],
        "bullet_point4": b[3], "bullet_point5": b[4],
        "product_description": (prod.get("content") or {}).get("description"),
        "recommended_browse_nodes": amazon_category(prod.get("product_name")),
        "color_name": a.get("primary_color"),
        "material_type": a.get("material"),
        "age_range_description": " - ".join(str(x) for x in (a.get("min_age"), a.get("max_age")) if x),
        "country_of_origin": src.get("country_of_origin"),
        "are_batteries_required": comp.get("batteries_required"),
        "item_weight": it.get("weight"),
        "item_weight_unit_of_measure": "pounds" if it.get("weight") else "",
        "item_length": it.get("length"), "item_width": it.get("width"),
        "item_height": it.get("height"),
        "item_length_unit_of_measure": "inches" if it.get("length") else "",
        "condition_type": "New",
    }


def select_products(vendor, ds_list):
    lc = load_lifecycles()
    slug = brand_slug(vendor) if vendor else None
    ds_set = set(str(d).strip() for d in ds_list) if ds_list else None
    out = []
    for p in sorted(glob.glob(os.path.join(PRODUCTS_DIR, "**", "*.json"), recursive=True)):
        d = json.load(open(p, encoding="utf-8"))
        if slug and brand_slug(d.get("vendor")) != slug:
            continue
        if ds_set and str(d.get("ds_number")).strip() not in ds_set:
            continue
        if lc.get(d.get("ds_number")) == "discontinued":
            continue
        out.append(d)
    return out


def write_generic(rows, out_path):
    wb = Workbook(); ws = wb.active; ws.title = "Amazon"
    ws.append(AMAZON_FIELDS)
    fill = PatternFill("solid", fgColor="2F5496")
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF"); c.fill = fill
        c.alignment = Alignment(vertical="center", wrap_text=True)
    for r in rows:
        ws.append([r.get(f, "") if r.get(f) is not None else "" for f in AMAZON_FIELDS])
    ws.freeze_panes = "A2"
    for i, f in enumerate(AMAZON_FIELDS, 1):
        ws.column_dimensions[get_column_letter(i)].width = 38 if ("name" in f or "bullet" in f or "description" in f or "image" in f or "browse" in f) else 16
    wb.save(out_path)


def fill_template(rows, template_path, out_path):
    wb = load_workbook(template_path)
    ws = wb["Template"] if "Template" in wb.sheetnames else wb[wb.sheetnames[0]]
    # find the field-name header row (the one containing 'item_sku')
    hdr_row = None
    for r in range(1, min(ws.max_row, 12) + 1):
        vals = [str(ws.cell(r, c).value).strip().lower() if ws.cell(r, c).value else "" for c in range(1, ws.max_column + 1)]
        if "item_sku" in vals:
            hdr_row = r
            colmap = {vals[c - 1]: c for c in range(1, ws.max_column + 1) if vals[c - 1]}
            break
    if not hdr_row:
        sys.exit("ERROR: couldn't find an 'item_sku' header row in the template's Template sheet.")
    start = ws.max_row + 1
    for i, r in enumerate(rows):
        for field, val in r.items():
            c = colmap.get(field)
            if c and val not in (None, ""):
                ws.cell(start + i, c, val)
    wb.save(out_path)
    return len([f for f in AMAZON_FIELDS if f in colmap])


def main():
    ap = argparse.ArgumentParser()
    sel = ap.add_mutually_exclusive_group(required=True)
    sel.add_argument("--vendor")
    sel.add_argument("--ds", nargs="+")
    ap.add_argument("--template", help="Amazon category flat-file to fill (else generic)")
    ap.add_argument("--out", required=True)
    args = ap.parse_args()

    prods = select_products(args.vendor, args.ds)
    if not prods:
        sys.exit("No matching (non-discontinued) products found.")
    rows = [row_for(p) for p in prods]

    if args.template:
        matched = fill_template(rows, args.template.strip('"'), args.out)
        print(f"Filled Amazon template with {len(rows)} items ({matched} fields mapped) -> {args.out}")
    else:
        write_generic(rows, args.out)
        print(f"Wrote generic Amazon flat-file ({len(rows)} items) -> {args.out}")


if __name__ == "__main__":
    main()
