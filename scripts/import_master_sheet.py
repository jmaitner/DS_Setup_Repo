#!/usr/bin/env python3
"""
import_master_sheet.py  —  Import a "Master Setup Data" consolidated export.

This is a DIFFERENT layout from the per-vendor DS Only template: headers are in ROW 1,
data starts ROW 2, columns are named (not positional), and there's a leading "Source File"
column. It's the format of "Master Product Setup Data.xlsx". This reader maps by fixed
column index (validated against that file's headers) into the same product JSON shape the
rest of the catalog uses, so validate/export/index/dashboard all work unchanged.

Usage:
  python scripts/import_master_sheet.py "C:\\path\\Master Product Setup Data.xlsx" --vendor "Bandai Namco"
  python scripts/import_master_sheet.py "<file>" --vendor "Bandai Namco" --dry-run
"""

import argparse
import json
import os
import sys
from datetime import date

for _stream in (sys.stdout, sys.stderr):
    try:
        _stream.reconfigure(encoding="utf-8")
    except (AttributeError, ValueError):
        pass

try:
    from openpyxl import load_workbook
except ImportError:
    sys.exit("ERROR: openpyxl not installed. Run: pip install -r requirements.txt")

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from ds_schema import _price, brand_slug

REPO_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
PRODUCTS_DIR = os.path.join(REPO_ROOT, "products")

_NA = {"", "n/a", "na", "none", "xxxx", "n/a "}


def t(v):
    """Text cell -> stripped string or None (N/A-style placeholders -> None)."""
    if v is None:
        return None
    s = str(v).strip()
    return None if s.lower() in _NA else s


def num(v):
    """Numeric cell -> float or None (handles N/A / blank / stray text)."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return v
    s = str(v).strip().replace(",", "")
    if s.lower() in _NA:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def build_product(ws, row, vendor, today):
    def C(c):
        return ws.cell(row=row, column=c).value

    feats = [t(C(c)) for c in range(51, 59)]      # Feature #1..#8
    feats = [f for f in feats if f]
    bt, bq = t(C(63)), t(C(64))                   # battery type, battery qty
    battery_type_qty = " x".join([x for x in (bt, bq) if x]) or None
    upc = t(C(4))

    return {
        "ds_number": str(t(C(2))) if t(C(2)) else None,
        "vendor": vendor,
        "brand": t(C(5)),
        "vendor_item_number": t(C(3)),
        "product_name": t(C(6)),
        "identity": {"id_type": "UPC" if upc else None, "upc": str(upc) if upc else None},
        "pricing": {
            "drop_ship_cost": _price(C(8)), "wholesale_cost": _price(C(9)),
            "ds_cost_domestic": _price(C(10)), "import_cost": None, "msrp": _price(C(7)),
        },
        "sourcing": {"fob_point": t(C(11)), "country_of_origin": t(C(32)),
                     "harmonized_code": t(C(35))},
        "content": {"bullets": feats[:8], "keywords": None, "description": t(C(50))},
        "attributes": {
            "material": t(C(40)), "num_pieces": num(C(41)), "whats_in_box": t(C(43)),
            "primary_color": t(C(44)), "secondary_color": t(C(45)),
            "min_age": t(C(46)), "max_age": t(C(47)), "gender": t(C(48)),
            "assembly_required": t(C(20)), "assembly_instructions": t(C(25)),
        },
        "images": {"availability": t(C(39)), "urls": []},
        "dimensions": {
            "item": {"weight": num(C(12)), "weight_unit": None, "length": num(C(13)),
                     "length_unit": None, "width": num(C(14)), "width_unit": None,
                     "height": num(C(15)), "height_unit": None},
            "package": {"weight": num(C(16)), "weight_unit": None, "length": num(C(17)),
                        "length_unit": None, "width": num(C(18)), "width_unit": None,
                        "height": num(C(19)), "height_unit": None},
            "master_case": {"case_qty": num(C(26)), "length": num(C(28)), "length_unit": None,
                            "width": num(C(29)), "width_unit": None, "height": num(C(30)),
                            "height_unit": None, "weight": num(C(27)), "weight_unit": None},
        },
        "compliance": {
            "choking_hazard": t(C(66)), "lead_phthalates": None,
            "warranty_included": t(C(37)), "warranty_desc": t(C(38)),
            "batteries_required": t(C(60)), "batteries_included": t(C(62)),
            "battery_cell_comp": t(C(61)), "battery_type_qty": battery_type_qty,
            "packaging_type": None, "compliance_cert": None, "doc": None, "sds": None,
            "sds_url": None, "cpsia": t(C(75)), "test_reports": None, "cpc": None,
            "product_pics": None, "instructions": None, "letter_of_compliance": None,
            "documents": [],
        },
        "_meta": {
            "source_file": t(C(1)), "source_sheet": "Master Setup Data",
            "imported_at": today, "supplier_id": None,
            "vendor_slug": brand_slug(vendor), "last_verified": None,
        },
    }


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("xlsx")
    ap.add_argument("--vendor", required=True, help='e.g. "Bandai Namco"')
    ap.add_argument("--sheet", default="Master Setup Data")
    ap.add_argument("--update", action="store_true")
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()

    path = args.xlsx.strip().strip('"').strip("'")
    if not os.path.isfile(path):
        sys.exit(f"ERROR: file not found -> {path}")
    wb = load_workbook(path, data_only=True)
    if args.sheet not in wb.sheetnames:
        sys.exit(f"ERROR: sheet '{args.sheet}' not found. Tabs: {wb.sheetnames}")
    ws = wb[args.sheet]

    slug = brand_slug(args.vendor)
    today = date.today().isoformat()
    vendor_dir = os.path.join(PRODUCTS_DIR, slug)
    print(f"Vendor: {args.vendor}  -> products/{slug}/   (sheet '{args.sheet}')")

    created, updated, skipped = [], [], []
    for row in range(2, ws.max_row + 1):
        if t(ws.cell(row=row, column=2).value) is None:   # no DS#
            continue
        prod = build_product(ws, row, args.vendor, today)
        ds = str(prod["ds_number"]).strip()
        out = os.path.join(vendor_dir, f"DS{ds}.json")
        rel = os.path.relpath(out, REPO_ROOT)
        exists = os.path.exists(out)
        if exists and not args.update:
            skipped.append(ds)
            print(f"  - skip   DS{ds} (exists; --update to overwrite)")
            continue
        if args.dry_run:
            (updated if exists else created).append(ds)
            print(f"  ~ would {'update' if exists else 'create'}  DS{ds}  {prod.get('product_name')}")
            continue
        os.makedirs(vendor_dir, exist_ok=True)
        with open(out, "w", encoding="utf-8") as f:
            json.dump(prod, f, indent=2, ensure_ascii=False, default=str)
            f.write("\n")
        (updated if exists else created).append(ds)
        print(f"  {'updated' if exists else 'created'}  DS{ds}  {prod.get('product_name')}")

    print(f"\nCreated: {len(created)}  Updated: {len(updated)}  Skipped: {len(skipped)}")
    print("Next: set_supplier, init_status, validate, build_index, build_dashboard.")


if __name__ == "__main__":
    main()
